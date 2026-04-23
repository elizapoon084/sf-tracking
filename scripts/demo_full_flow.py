# -*- coding: utf-8 -*-
"""
demo_full_flow.py  —  完整流程示範 (Steps 1-13)
================================================
獨立腳本，不依賴任何現有模組。
全程用 JS injection 操作。

執行後彈出 10 行輸入框，每行貼入一個 WhatsApp 訂單即可批量自動完成全流程。
訂單格式範例：
  1084043-3件，1084065-2件，寄黃业偉 18125989028 廣東省深圳市龍崗區南聯劉屋村南段74號1樓
"""
import os
import sys
import time
import re
import json
import tkinter as tk
from tkinter import messagebox, ttk
from datetime import date
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

from playwright.sync_api import sync_playwright

# ─── 系統路徑 & URL (唔需要改) ────────────────────────────────────────────────

CHROME_PROFILE  = r"C:\ChromeAutomation"
IMAGES_DIR      = r"C:\Users\user\Desktop\順丰E順递\Images"
LOGS_DIR        = r"C:\Users\user\Desktop\順丰E順递\logs"
PRODUCTS_JSON   = r"C:\Users\user\Desktop\順丰E順递\data\products.json"
MONTHLY_ACCOUNT = "8526937071"

POS_URL     = "https://online-store-99126206.web.app/"
POS_PASS    = "0000"
VIP_PASS    = "941196"

SF_URL      = "https://hk.sf-express.com/hk/tc/ship/home"
SENDER_TEXT = "潘正儀 66832382 香港九龍新蒲崗大有街33號佳力工業大廈603室"

EXCEL_PATH  = r"C:\Users\user\Desktop\順丰E順递\data\tracking.xlsx"

_NUM_ROWS = 10

# ══════════════════════════════════════════════════════════════════════════════
# 解析工具
# ══════════════════════════════════════════════════════════════════════════════

def _load_products() -> dict:
    try:
        with open(PRODUCTS_JSON, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _parse_order(raw: str) -> dict:
    """解析 WhatsApp 訂單文字，返回 order dict。失敗則 raise ValueError。"""
    products = _load_products()

    items_pos = []
    items_sf  = []
    for m in re.finditer(r"(\d{5,10})[\s\-×xX*]+(\d+)\s*件", raw):
        sku, qty = m.group(1), int(m.group(2))
        prod = products.get(sku, {})
        items_pos.append({"sku": sku, "qty": qty})
        items_sf.append({
            "sku":        sku,
            "name":       prod.get("name", sku),
            "brand":      prod.get("brand", ""),
            "material":   prod.get("material", ""),
            "spec":       prod.get("spec", ""),
            "unit_price": float(prod.get("vip_price", 0)),
            "qty":        qty,
        })

    if not items_pos:
        raise ValueError("找唔到貨品 — 格式：貨號-數量件")

    m = re.search(r"寄\s*([^\s\d，,、]{1,10})\s+", raw)
    if m:
        name = m.group(1).strip("，,、 \t")
    else:
        # 冇「寄」字：名字喺最後一個貨品之後
        last_end = 0
        for im in re.finditer(r"(\d{5,10})[\s\-×xX*]+(\d+)\s*件", raw):
            last_end = im.end()
        remainder = raw[last_end:].strip().lstrip("，,、 \t")
        nm = re.match(r"([^\s\d，,、]{1,10})\s+", remainder)
        if not nm:
            raise ValueError("找唔到收件人名 — 格式：寄[姓名] 電話 地址")
        name = nm.group(1).strip("，,、 \t")

    m = re.search(r"1[3-9]\d{9}|[2-9]\d{7}", raw)
    if not m:
        raise ValueError("找唔到電話號碼")
    phone = m.group(0)

    m = re.search(r"(?:1[3-9]\d{9}|[2-9]\d{7})\s*(.+)", raw, re.DOTALL)
    if not m:
        raise ValueError("找唔到地址")
    address = re.split(r"[。\n]", m.group(1).strip())[0].strip()

    return {
        "name": name, "phone": phone, "address": address,
        "items_pos": items_pos, "items_sf": items_sf,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Step 1 — 10 行輸入框
# ══════════════════════════════════════════════════════════════════════════════

def show_order_input() -> list:
    """彈出 10 行訂單輸入框，解析後返回 [order, ...] list。取消則退出程式。"""
    results = [None]

    root = tk.Tk()
    root.title("順丰寄件 — 輸入客人訂單")
    root.resizable(True, True)
    root.geometry("860x620")

    # ── 10-row input grid ──────────────────────────────────────────────────
    input_lf = tk.LabelFrame(
        root,
        text="WhatsApp 訂單 (每行一個客人，貼入後按解析)",
        padx=6, pady=6,
    )
    input_lf.pack(fill="x", padx=12, pady=(12, 4))

    row_vars   = []
    row_status = []

    for i in range(_NUM_ROWS):
        sv  = tk.StringVar()
        rst = tk.StringVar(value="")
        row_vars.append(sv)
        row_status.append(rst)

        row_frame = tk.Frame(input_lf)
        row_frame.pack(fill="x", pady=1)

        tk.Label(row_frame, text=f"{i+1:2d}.", width=3,
                 anchor="e", font=("Courier", 10)).pack(side="left")
        tk.Entry(row_frame, textvariable=sv,
                 font=("", 10), relief="solid", bd=1).pack(
            side="left", fill="x", expand=True, padx=(4, 4))
        tk.Label(row_frame, textvariable=rst,
                 width=10, anchor="w", font=("", 9)).pack(side="left")

    btn_row_top = tk.Frame(input_lf)
    btn_row_top.pack(fill="x", pady=(6, 0))
    tk.Button(btn_row_top, text="🗑 清除全部",
              command=lambda: [sv.set("") or rst.set("")
                               for sv, rst in zip(row_vars, row_status)],
              bg="#95a5a6", fg="white", padx=8, pady=3).pack(side="left")

    # ── Parsed results table ───────────────────────────────────────────────
    result_lf = tk.LabelFrame(root, text="已解析訂單", padx=6, pady=4)
    result_lf.pack(fill="both", expand=True, padx=12, pady=4)

    cols = ("#", "收件人", "電話", "貨品", "狀態")
    tree = ttk.Treeview(result_lf, columns=cols, show="headings", height=8)
    tree.heading("#",    text="#");    tree.column("#",    width=30, anchor="center")
    tree.heading("收件人", text="收件人"); tree.column("收件人", width=90)
    tree.heading("電話",  text="電話");  tree.column("電話",  width=120)
    tree.heading("貨品",  text="貨品");  tree.column("貨品",  width=380)
    tree.heading("狀態",  text="狀態");  tree.column("狀態",  width=100, anchor="center")
    tree.pack(fill="both", expand=True)
    tree.tag_configure("ok",    foreground="#27ae60")
    tree.tag_configure("error", foreground="#e74c3c")

    parsed_orders = []

    def on_parse():
        try:
            parsed_orders.clear()
            for item in tree.get_children():
                tree.delete(item)

            ok_count = 0
            for i, sv in enumerate(row_vars):
                raw = sv.get().strip()
                if not raw:
                    row_status[i].set("")
                    continue
                try:
                    order = _parse_order(raw)
                    items_str = ", ".join(
                        f"{it['name']}×{it['qty']}"
                        for it in order["items_sf"])
                    parsed_orders.append(order)
                    tree.insert("", "end",
                                values=(i+1, order["name"], order["phone"],
                                        items_str, "待寄出"),
                                tags=("ok",))
                    row_status[i].set("✅")
                    ok_count += 1
                except ValueError as e:
                    row_status[i].set("❌")
                    tree.insert("", "end",
                                values=(i+1, "—", "—", str(e)[:60], "解析失敗"),
                                tags=("error",))

            if ok_count:
                run_btn.config(state="normal")
                info_var.set(f"✅ 解析完成：{ok_count} 個訂單準備好，按「開始自動化」")
            else:
                run_btn.config(state="disabled")
                info_var.set("❌ 沒有成功解析的訂單")

        except Exception as e:
            messagebox.showerror("解析錯誤", f"內部錯誤：{e}")

    def on_run():
        if parsed_orders:
            results[0] = list(parsed_orders)
            root.destroy()

    def on_cancel():
        sys.exit(0)

    # ── Info label ─────────────────────────────────────────────────────────
    info_var = tk.StringVar(value="（貼入訂單後按「解析全部訂單」）")
    tk.Label(root, textvariable=info_var, anchor="w",
             fg="#555", font=("", 9)).pack(fill="x", padx=14)

    # ── Bottom buttons ─────────────────────────────────────────────────────
    btn_row = tk.Frame(root)
    btn_row.pack(pady=(4, 14))

    tk.Button(btn_row, text="🔍 解析全部訂單",
              command=on_parse,
              bg="#2980b9", fg="white",
              padx=14, pady=6, font=("", 11)).pack(side="left", padx=6)

    run_btn = tk.Button(btn_row, text="🚀 開始自動化",
                        command=on_run,
                        bg="#27ae60", fg="white",
                        padx=14, pady=6, font=("", 11, "bold"),
                        state="disabled")
    run_btn.pack(side="left", padx=6)

    tk.Button(btn_row, text="✖ 取消",
              command=on_cancel,
              bg="#e74c3c", fg="white",
              padx=14, pady=6, font=("", 11)).pack(side="left", padx=6)

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()
    return results[0]


# ══════════════════════════════════════════════════════════════════════════════
# 取得訂單資料
# ══════════════════════════════════════════════════════════════════════════════

orders = show_order_input()   # list of order dicts

# ─── 關閉殘留 Chrome 進程 & 清理 lock files ──────────────────────────────────────

# 只 kill 用 ChromeAutomation profile 的 Chrome，唔影響其他 Chrome 視窗
subprocess.run(
    ["powershell", "-Command",
     "Get-WmiObject Win32_Process | Where-Object { $_.CommandLine -like '*ChromeAutomation*' } | ForEach-Object { $_.Terminate() }"],
    capture_output=True
)
time.sleep(1.5)

for lf in ["lockfile", "SingletonLock", "SingletonSocket", "SingletonCookie"]:
    try:
        os.remove(os.path.join(CHROME_PROFILE, lf))
    except Exception:
        pass

os.makedirs(LOGS_DIR, exist_ok=True)
today = date.today().strftime("%Y%m%d")


# ══════════════════════════════════════════════════════════════════════════════
# 工具函數
# ══════════════════════════════════════════════════════════════════════════════

def shot(page, label):
    p = os.path.join(LOGS_DIR, f"demo_{label}.png")
    page.screenshot(path=p, full_page=False)
    print(f"  📸 {p}")


def js_click_text(page, *texts, min_w=20, min_h=10):
    js_texts = str(list(texts))
    return page.evaluate(f"""() => {{
        const targets = {js_texts};
        for (const el of document.querySelectorAll('*')) {{
            if (el.offsetParent === null || el.children.length > 0) continue;
            const t = el.textContent.trim();
            if (!targets.includes(t)) continue;
            const r = el.getBoundingClientRect();
            if (r.width < {min_w} || r.height < {min_h}) continue;
            el.click();
            return {{ok: true, text: t}};
        }}
        return {{ok: false}};
    }}""")


def js_fill_by_label(page, label, value):
    info = page.evaluate(f"""() => {{
        const targets = ['{label}', '{label}：', '{label}:'];
        const labels = [];
        for (const el of document.querySelectorAll('*')) {{
            if (el.offsetParent === null) continue;
            const t = el.textContent.trim();
            if (targets.includes(t)) {{ labels.push(el); continue; }}
            if (t.length < 12 && t.includes('{label}')) labels.push(el);
        }}
        if (!labels.length) return null;
        labels.sort((a,b) => a.children.length - b.children.length);
        for (const lbl of labels) {{
            let p = lbl.parentElement;
            for (let d = 0; d < 5; d++) {{
                if (!p) break;
                for (const inp of p.querySelectorAll('input:not([type=hidden]):not([type=radio]):not([type=checkbox])')) {{
                    if (inp.offsetParent === null) continue;
                    const r = inp.getBoundingClientRect();
                    if (r.width === 0) continue;
                    return {{x: r.left + r.width/2, y: r.top + r.height/2}};
                }}
                p = p.parentElement;
            }}
        }}
        return null;
    }}""")
    if not info:
        print(f"  ⚠️  找不到欄位: {label}")
        return False
    page.mouse.click(info["x"], info["y"])
    time.sleep(0.3)
    page.keyboard.press("Control+A")
    page.keyboard.press("Delete")
    page.keyboard.type(str(value), delay=30)
    time.sleep(0.3)
    return True


def smart_fill(page, text, which):
    page.locator("span:has-text('智慧填寫')").nth(which).click()
    time.sleep(0.8)
    ta = page.locator("textarea[class*='intelAddr'], textarea[placeholder*='陳先生']").first
    ta.wait_for(state="visible", timeout=5000)
    ta.click()
    ta.type(text, delay=20)
    time.sleep(0.5)
    page.evaluate("""() => {
        for (const el of document.querySelectorAll('*')) {
            if (el.childNodes.length===1 && el.firstChild.nodeType===3
                && el.firstChild.textContent.trim()==='識別') {
                el.click(); return;
            }
        }
    }""")
    time.sleep(2.5)


def wait_dialog_closed(page, timeout=6):
    for _ in range(timeout * 2):
        gone = page.evaluate("""() =>
            document.querySelectorAll('[role="dialog"][data-state="open"]').length === 0
        """)
        if gone:
            return True
        time.sleep(0.5)
    return False


# ══════════════════════════════════════════════════════════════════════════════
# Excel 追蹤表工具
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_HEADERS = [
    "日期", "客人名", "POS訂單號", "順丰運單號",
    "收件人", "收件電話", "收件地址",
    "貨品摘要", "件數", "VIP總額(HKD)",
    "付款方式", "運費(HKD)", "最新狀態",
    "狀態更新時間", "異常標記", "小票檔案路徑", "備註", "稅金(HKD)",
]
EXCEL_SHEET = "追蹤表"

_HDR_COL = {h: i+1 for i, h in enumerate(EXCEL_HEADERS)}   # header → col index

def _ensure_excel():
    """建立 Excel 檔案（若不存在），設定標題行格式。"""
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET

    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx, hdr in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border

    # 欄寬
    col_widths = {
        "日期": 12, "客人名": 12, "POS訂單號": 14, "順丰運單號": 20,
        "收件人": 12, "收件電話": 14, "收件地址": 36,
        "貨品摘要": 34, "件數": 7, "VIP總額(HKD)": 13,
        "付款方式": 12, "運費(HKD)": 12, "最新狀態": 14,
        "狀態更新時間": 20, "異常標記": 10, "小票檔案路徑": 40,
        "備註": 16, "稅金(HKD)": 12,
    }
    for hdr, width in col_widths.items():
        ws.column_dimensions[get_column_letter(_HDR_COL[hdr])].width = width

    ws.freeze_panes = "A2"          # 鎖定標題行
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"  📊 Excel 表已建立: {EXCEL_PATH}")


def append_order_to_excel(order: dict, waybill: str, pdf_path: str, pos_order_no: str):
    """
    把一個訂單寫入 Excel 追蹤表（新增一行）。
    order keys: name, phone, items_sf
    """
    _ensure_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[EXCEL_SHEET]

    items_str  = " / ".join(
        f"{it['name']}×{it['qty']}" for it in order.get("items_sf", []))
    total_amt  = sum(
        it.get("unit_price", 0) * it.get("qty", 1)
        for it in order.get("items_sf", []))
    total_qty  = sum(it.get("qty", 1) for it in order.get("items_sf", []))

    row_data = {
        "日期":          today,
        "客人名":        order.get("name", ""),
        "POS訂單號":     pos_order_no,
        "順丰運單號":    waybill,
        "收件人":        order.get("name", ""),
        "收件電話":      order.get("phone", ""),
        "收件地址":      order.get("address", ""),
        "貨品摘要":      items_str,
        "件數":          total_qty,
        "VIP總額(HKD)":  round(total_amt, 1),
        "付款方式":      "月結",
        "最新狀態":      "待更新",
        "小票檔案路徑":  pdf_path,
    }

    next_row = ws.max_row + 1
    data_fill   = PatternFill("solid", fgColor="EBF3FB")
    alt_fill    = PatternFill("solid", fgColor="FFFFFF")
    fill        = data_fill if next_row % 2 == 0 else alt_fill
    thin_side   = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)

    for hdr, col in _HDR_COL.items():
        cell = ws.cell(row=next_row, column=col, value=row_data.get(hdr, ""))
        cell.fill   = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    wb.save(EXCEL_PATH)
    print(f"  📊 Excel 已記錄: Row {next_row}  {order['name']}  {waybill}")


# ══════════════════════════════════════════════════════════════════════════════
# 主流程：逐個訂單執行
# ══════════════════════════════════════════════════════════════════════════════

success_count = 0
failed_count  = 0
completed_orders = []   # [(customer, waybill, pdf_path)]

with sync_playwright() as pw:
    ctx = pw.chromium.launch_persistent_context(
        CHROME_PROFILE, channel="chrome", headless=False,
        args=["--disable-blink-features=AutomationControlled"],
        slow_mo=150, viewport={"width": 1280, "height": 900},
    )

    for order_idx, order in enumerate(orders):
        DEMO_CUSTOMER  = order["name"]
        DEMO_PHONE     = order["phone"]
        DEMO_ADDRESS   = order["address"]
        DEMO_POS_ITEMS = order["items_pos"]
        DEMO_SF_ITEMS  = order["items_sf"]

        print(f"\n{'='*60}")
        print(f"  訂單 {order_idx+1}/{len(orders)}: {DEMO_CUSTOMER}")
        print(f"{'='*60}")

        try:
            # ════════════════════════════════════════════════════════════════
            # 階段一：POS 打小票 (Steps 1-5)
            # ════════════════════════════════════════════════════════════════

            print("\n▶ Step 2: 登入 POS 後台")
            pos_page = ctx.new_page()
            pos_page.goto(POS_URL, wait_until="domcontentloaded", timeout=20000)
            time.sleep(3)

            pos_page.locator("button:has-text('后台管理')").first.click()
            time.sleep(0.8)
            pos_page.locator("input[type='password']").first.fill(POS_PASS)
            pos_page.keyboard.press("Enter")
            time.sleep(1.5)
            print("  ✅ 後台管理已登入")

            print("\n▶ Step 3: 啟動 VIP 價")
            pos_page.locator("button:has-text('VIP價')").first.click()
            time.sleep(0.8)
            pos_page.locator("input[type='password']").first.fill(VIP_PASS)
            pos_page.keyboard.press("Enter")
            time.sleep(1.5)
            print("  ✅ VIP 價已啟動")

            print(f"\n▶ Step 3: 加入 {len(DEMO_POS_ITEMS)} 種貨品")
            for item in DEMO_POS_ITEMS:
                btn = pos_page.locator(f"button:has-text('{item['sku']}')").first
                btn.wait_for(state="visible", timeout=5000)
                for _ in range(item["qty"]):
                    btn.click()
                    time.sleep(0.25)
                print(f"  ✅ SKU {item['sku']} × {item['qty']}")

            shot(pos_page, f"{order_idx+1:02d}_01_pos_cart")

            print("\n▶ Step 4: 結帳 (VIP 價)")
            pos_page.locator("button:has-text('結帳')").first.click()
            time.sleep(1.5)

            # 現金係 default，直接按確認出小票
            pos_page.locator("button:has-text('確認，出小票')").first.click()
            time.sleep(2.5)
            shot(pos_page, f"{order_idx+1:02d}_02_pos_receipt")

            body_text = pos_page.inner_text("body")
            m = re.search(r"ORD-\d+", body_text)
            pos_order_no = m.group(0) if m else f"ORD-{today}"
            print(f"  ✅ POS 訂單號: {pos_order_no}")

            print("\n▶ Step 5: 儲存小票到客人 Folder")
            save_dir = os.path.join(IMAGES_DIR, DEMO_CUSTOMER)
            os.makedirs(save_dir, exist_ok=True)
            file_base = f"{today}_{DEMO_CUSTOMER}_{pos_order_no}"
            pdf_path  = os.path.join(save_dir, file_base + ".pdf")
            png_path  = os.path.join(save_dir, file_base + ".png")

            # 按藍色 DOWNLOAD 按鈕，Playwright 攔截下載並儲存到客人 folder
            try:
                with pos_page.expect_download(timeout=15000) as dl_info:
                    pos_page.evaluate("""() => {
                        const cands = ['DOWNLOAD','Download','download','下載','下載小票','PDF','列印','Print'];
                        for (const t of cands) {
                            for (const el of document.querySelectorAll('button, a, [role="button"]')) {
                                if (el.offsetParent === null) continue;
                                const txt = el.textContent.trim();
                                if (txt === t || txt.includes(t)) { el.click(); return; }
                            }
                        }
                    }""")
                dl = dl_info.value
                dl.save_as(pdf_path)
                print(f"  ✅ 小票已儲存: {pdf_path}")
            except Exception as dl_err:
                print(f"  ⚠️  下載攔截失敗，改截圖備份: {dl_err}")
                pos_page.screenshot(path=png_path, full_page=False)
                pdf_path = png_path
                print(f"  ✅ 截圖備份: {png_path}")

            try:
                done_btn = pos_page.locator("button:has-text('完成')").first
                if done_btn.is_visible(timeout=2000):
                    done_btn.click()
                    time.sleep(0.8)
            except Exception:
                pass
            pos_page.close()

            print("\n  階段二：包裝 (Step 6) — 人手執行")

            # ── 確認框：等待人手執行包裝後才繼續去順豐 ──────────────────────
            proceed = messagebox.askyesno(
                "Step 6 — 包裝確認",
                f"客人：{DEMO_CUSTOMER}\n"
                f"小票：{pdf_path}\n\n"
                "貨品已執好、打包完成？\n\n"
                "按「是」繼續去順豐落單\n"
                "按「否」跳過此訂單",
            )
            if not proceed:
                print(f"  ⏭️  {DEMO_CUSTOMER} 已跳過順豐落單")
                completed_orders.append((DEMO_CUSTOMER, "跳過", pdf_path))
                success_count += 1
                continue

            # ════════════════════════════════════════════════════════════════
            # 階段三：填順丰快遞單 (Steps 7-13)
            # ════════════════════════════════════════════════════════════════

            print("\n▶ Step 7: 開啟順豐寄件頁面")
            sf_page = ctx.new_page()
            sf_page.goto(SF_URL, wait_until="domcontentloaded", timeout=20000)
            time.sleep(3)

            print("\n▶ Step 8: 寄件人 → 自寄")
            smart_fill(sf_page, SENDER_TEXT, 0)

            sf_page.evaluate("""() => {
                const inputs = document.querySelectorAll("input[name='contactName']");
                if (inputs[0]) {
                    const setter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype,'value').set;
                    setter.call(inputs[0],'潘正儀');
                    inputs[0].dispatchEvent(new Event('input',{bubbles:true}));
                    inputs[0].dispatchEvent(new Event('change',{bubbles:true}));
                }
            }""")
            time.sleep(0.5)

            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    if (el.textContent.trim()==='自寄' && el.offsetParent!==null) {
                        const r = el.getBoundingClientRect();
                        if (r.width>40 && r.height>20) { el.click(); return; }
                    }
                }
            }""")
            time.sleep(1.5)
            print("  ✅ 自寄已選")

            print("\n▶ Step 9: 收件人智慧填寫")
            recip_text = f"{DEMO_CUSTOMER} {DEMO_PHONE} {DEMO_ADDRESS}"
            smart_fill(sf_page, recip_text, 1)
            shot(sf_page, f"{order_idx+1:02d}_03_sf_recipient")
            print("  ✅ 收件人已填")

            print(f"\n▶ Step 10: 填 {len(DEMO_SF_ITEMS)} 件物品信息")

            for idx, item in enumerate(DEMO_SF_ITEMS):
                print(f"\n  ── 物品 {idx+1}: {item['name']} ──")

                info = sf_page.evaluate("""() => {
                    const candidates = ['+新增物品','新增物品','+ 新增物品'];
                    for (const el of document.querySelectorAll('*')) {
                        const t = el.textContent.trim();
                        if (!candidates.includes(t)) continue;
                        if (el.offsetParent===null || el.children.length>0) continue;
                        const r = el.getBoundingClientRect();
                        if (r.width<20||r.height<10) continue;
                        return {absY: r.top+window.scrollY, h:r.height};
                    }
                    return null;
                }""")
                if info:
                    sf_page.evaluate(f"window.scrollTo(0,{info['absY']}-450+{info['h']}/2)")
                    time.sleep(0.8)
                click_info = sf_page.evaluate("""() => {
                    const candidates = ['+新增物品','新增物品','+ 新增物品'];
                    for (const el of document.querySelectorAll('*')) {
                        const t = el.textContent.trim();
                        if (!candidates.includes(t)) continue;
                        if (el.offsetParent===null || el.children.length>0) continue;
                        const r = el.getBoundingClientRect();
                        if (r.width<20||r.height<10) continue;
                        return {x:r.left+r.width/2, y:r.top+r.height/2};
                    }
                    return null;
                }""")
                if not click_info:
                    print("  ❌ 搵唔到 +新增物品")
                    break
                sf_page.mouse.click(click_info["x"], click_info["y"])
                time.sleep(2)

                sf_page.evaluate("""() => {
                    for (const el of document.querySelectorAll('*')) {
                        if (el.textContent.trim()==='物品' && el.offsetParent!==null
                            && el.children.length===0) {
                            const r = el.getBoundingClientRect();
                            if (r.width>10&&r.height>5) { el.click(); return; }
                        }
                    }
                }""")
                time.sleep(1.5)

                js_fill_by_label(sf_page, "物品名稱", item["name"])
                sf_page.keyboard.press("Tab")
                time.sleep(4)
                print("  ✅ 物品名稱")

                js_fill_by_label(sf_page, "品牌", item["brand"]); time.sleep(2)
                print("  ✅ 品牌")
                js_fill_by_label(sf_page, "材質", item["material"]); time.sleep(2)
                print("  ✅ 材質")
                js_fill_by_label(sf_page, "規格型號", item["spec"]); time.sleep(2)
                print("  ✅ 規格")
                js_fill_by_label(sf_page, "物品單價", item["unit_price"]); time.sleep(0.5)
                print("  ✅ 單價")
                js_fill_by_label(sf_page, "物品數量", item["qty"]); time.sleep(0.5)
                print("  ✅ 數量")

                sf_page.evaluate("""() => {
                    for (const el of document.querySelectorAll('[class*="package-declaration_confirm"]')) {
                        if (el.offsetParent===null) continue;
                        if (el.textContent.trim()!=='確認') continue;
                        const r = el.getBoundingClientRect();
                        if (r.width<30||r.height<15) continue;
                        el.click(); return;
                    }
                }""")
                time.sleep(2)
                wait_dialog_closed(sf_page)
                print("  ✅ 已確認")

            print("\n▶ Step 11: 付款方式 → 寄付月結")
            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    if (el.offsetParent===null || el.children.length>0) continue;
                    const t = el.textContent.trim();
                    if (t!=='月結' && t!=='寄付月結') continue;
                    const r = el.getBoundingClientRect();
                    if (r.width<20||r.height<15) continue;
                    el.click(); return;
                }
            }""")
            time.sleep(1.5)
            print("  ✅ 寄付月結已選")

            print(f"\n▶ Step 12: 填月結卡號 {MONTHLY_ACCOUNT}")
            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    if (el.textContent.trim()==='付款方式' && el.offsetParent!==null)
                        { el.scrollIntoView({block:'center'}); return; }
                }
            }""")
            time.sleep(0.8)
            acct_info = sf_page.evaluate("""() => {
                for (const inp of document.querySelectorAll('input')) {
                    if (inp.offsetParent===null) continue;
                    const ph = inp.placeholder||'';
                    const r = inp.getBoundingClientRect();
                    if (r.width===0||r.height===0) continue;
                    if (ph.includes('月結')||ph.includes('卡號'))
                        return {x:r.left+r.width/2, y:r.top+r.height/2};
                }
                return null;
            }""")
            if acct_info:
                sf_page.mouse.click(acct_info["x"], acct_info["y"])
                time.sleep(0.3)
                sf_page.mouse.click(acct_info["x"], acct_info["y"])
                time.sleep(0.3)
                sf_page.keyboard.press("Control+A")
                sf_page.keyboard.press("Delete")
                sf_page.keyboard.type(MONTHLY_ACCOUNT, delay=50)
                time.sleep(0.8)
                sf_page.evaluate(f"""() => {{
                    for (const el of document.querySelectorAll('*')) {{
                        if (el.offsetParent===null||el.children.length>0) continue;
                        const t = el.textContent.trim();
                        if (t.includes('{MONTHLY_ACCOUNT}')&&t.length<50) {{
                            const r=el.getBoundingClientRect();
                            if (r.width<30) return;
                            el.click(); return;
                        }}
                    }}
                }}""")
                sf_page.keyboard.press("Tab")
                time.sleep(0.5)
                print("  ✅ 月結卡號已填")
            else:
                print("  ⚠️  月結卡號欄位未找到")

            print("\n▶ Step 13a: 閱讀並同意")
            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('[class*="agreedCheckbox"]')) {
                    if (el.offsetParent!==null)
                        { el.scrollIntoView({block:'center'}); return; }
                }
            }""")
            time.sleep(0.8)
            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('[class*="agreedCheckbox"]')) {
                    if (el.offsetParent===null) continue;
                    const r = el.getBoundingClientRect();
                    if (r.width<30) continue;
                    el.click(); return;
                }
                for (const el of document.querySelectorAll('[class*="checkbox_checkbox"]')) {
                    if (el.offsetParent===null) continue;
                    const r = el.getBoundingClientRect();
                    if (r.width<30) continue;
                    const txt = (el.parentElement||el).textContent||'';
                    if (txt.includes('閱讀')&&txt.includes('同意'))
                        { el.click(); return; }
                }
            }""")
            time.sleep(2)
            print("  ✅ Checkbox 已勾")

            sf_page.evaluate("""() => {
                const cands=['同意本條款,下次不再提示','同意本條款，下次不再提示',
                             '同意本條款','同意並繼續','同意'];
                for (const t of cands) {
                    for (const el of document.querySelectorAll('*')) {
                        if (el.offsetParent===null) continue;
                        if (el.textContent.trim()!==t) continue;
                        if (el.children.length>0) continue;
                        const r=el.getBoundingClientRect();
                        if (r.width<50||r.height<20) continue;
                        el.click(); return;
                    }
                }
            }""")
            time.sleep(1.5)

            print("\n▶ Step 13: 下單")
            sf_page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(0.5)

            waybill = "未擷取"
            submit_result = {}
            for attempt in range(8):
                submit_result = sf_page.evaluate("""() => {
                    for (const el of document.querySelectorAll('[class*="submitBtn"]')) {
                        if (el.offsetParent===null) continue;
                        if (!el.textContent.trim().includes('下單')) continue;
                        const cls=(el.className||'').toString();
                        if (cls.includes('disabled')) return {ok:false,reason:'disabled'};
                        const r=el.getBoundingClientRect();
                        if (r.width<40) continue;
                        el.click();
                        return {ok:true};
                    }
                    for (const el of document.querySelectorAll('[role="button"]')) {
                        if (el.offsetParent===null) continue;
                        if (el.textContent.trim()!=='下單') continue;
                        const cls=(el.className||'').toString();
                        if (cls.includes('disabled')) return {ok:false,reason:'role_disabled'};
                        el.click();
                        return {ok:true};
                    }
                    return {ok:false,reason:'not_found'};
                }""")
                if submit_result.get("ok"):
                    print(f"  ✅ 下單成功 (attempt {attempt+1})")
                    break
                time.sleep(2)
            else:
                print(f"  ❌ 下單失敗: {submit_result}")

            time.sleep(3)
            shot(sf_page, f"{order_idx+1:02d}_04_sf_submitted")

            for _ in range(5):
                content = sf_page.content()
                m = re.search(r"SF\d{10,}", content)
                if m:
                    waybill = m.group(0)
                    break
                time.sleep(1)

            print(f"\n  ✅ {DEMO_CUSTOMER} 完成！運單號: {waybill}")
            print(f"  📄 小票: {pdf_path}")

            # ── 寫入 Excel 追蹤表 ────────────────────────────────────────────
            append_order_to_excel(order, waybill, pdf_path, pos_order_no)

            sf_page.close()
            success_count += 1
            completed_orders.append((DEMO_CUSTOMER, waybill, pdf_path))

        except Exception as e:
            print(f"\n  ❌ {DEMO_CUSTOMER} 失敗: {e}")
            failed_count += 1
            try:
                sf_page.close()
            except Exception:
                pass

    # ── 完成摘要 ────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  批量完成：✅ {success_count} 成功  ❌ {failed_count} 失敗")
    print(f"\n  本次運單號：")
    for customer, waybill, pdf in completed_orders:
        print(f"    {customer}  →  {waybill}")
    print(f"\n  📊 Excel 追蹤表：{EXCEL_PATH}")
    print(f"{'='*60}")

    # ── 查件：開順豐追蹤頁面 ──────────────────────────────────────────────
    if completed_orders:
        waybill_list = ",".join(w for _, w, _ in completed_orders if w != "未擷取")
        if waybill_list:
            track_url = (
                f"https://www.sf-express.com/cn/sc/dynamic_function/waybill/"
                f"#search/bill-number/{waybill_list}"
            )
            print(f"\n  正在開啟順豐查件頁面…")
            # 在現有瀏覽器開新頁查件
            track_page = ctx.new_page()
            track_page.goto(track_url, wait_until="domcontentloaded", timeout=15000)

    input("\n按 Enter 關閉瀏覽器…")
    ctx.close()
