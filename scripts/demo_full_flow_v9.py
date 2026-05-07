# -*- coding: utf-8 -*-
"""
demo_full_flow.py  —  完整流程示範 (Steps 1-13)
================================================
獨立腳本，不依賴任何現有模組。
全程用 JS injection 操作。

執行後彈出 10 行輸入框，每行貼入一個 WhatsApp 訂單即可批量自動完成全流程。
訂單格式範例：
  1084043-3件，1084065-2件，寄黃业偉 18125989028 廣東省深圳市龍崗區南聯劉屋村南段74號1樓
"""
import os
import sys
import time
import re
import json
import base64
import tkinter as tk
from tkinter import messagebox, ttk
from datetime import date
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

from playwright.sync_api import sync_playwright

# ─── 系統路徑 & URL (唔需要改) ────────────────────────────────────────────────

CHROME_PROFILE  = r"C:\ChromeAutomation"
IMAGES_DIR      = r"C:\Users\user\Desktop\順丰E順递\Images"   # local backup only
ORDERS_DIR      = r"C:\Users\user\Desktop\順丰E順递\data\orders"  # git-tracked → Streamlit Cloud
LOGS_DIR        = r"C:\Users\user\Desktop\順丰E順递\logs"
PRODUCTS_JSON   = r"C:\Users\user\Desktop\順丰E順递\data\products.json"
MONTHLY_ACCOUNT = "8526937071"

POS_URL     = "https://online-store-99126206.web.app/"
POS_PASS    = "0000"
VIP_PASS    = "941196"

SF_URL         = "https://hk.sf-express.com/hk/tc/ship/home"
SF_LIST_URL    = "https://hk.sf-express.com/hk/tc/waybill/list"
SF_DETAIL_BASE = "https://hk.sf-express.com/hk/tc/waybill/appointment-detail"
SENDER_TEXT = "潘正儀 66832382 香港九龍新蒲崗大有街33號佳力工業大廈603室"

EXCEL_PATH  = r"C:\Users\user\Desktop\順丰E順递\data\tracking.xlsx"

_NUM_ROWS = 20

# ══════════════════════════════════════════════════════════════════════════════
# 解析工具
# ══════════════════════════════════════════════════════════════════════════════

def _load_products() -> dict:
    try:
        with open(PRODUCTS_JSON, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _parse_order(raw: str) -> dict:
    """解析 WhatsApp 訂單文字，返回 order dict。失敗則 raise ValueError。"""
    products = _load_products()

    items_pos = []
    items_sf  = []
    ITEM_RE = re.compile(
        r"(\d{5,10})"
        r"[^\d]*?"
        r"(\d{1,4})"
        r"\s*[件個盒包條罐支瓶箱套份粒]"
    )
    last_end = 0

    for m in ITEM_RE.finditer(raw):
        sku, qty = m.group(1), int(m.group(2))
        last_end = m.end()
        prod = products.get(sku, {})
        items_pos.append({"sku": sku, "qty": qty})
        items_sf.append({
            "sku":        sku,
            "name":       prod.get("name", sku),
            "brand":      prod.get("brand", ""),
            "material":   prod.get("material", ""),
            "spec":       prod.get("spec", ""),
            "unit_price": float(prod.get("vip_price", 0)),
            "qty":        qty,
        })

    if not items_pos:
        raise ValueError("找唔到貨品 — 請用：型號 數量件/個/盒")

    nm = re.search(r"寄\s*([^\s\d，,、]{1,10})\s+", raw)
    if nm:
        name = nm.group(1).strip("，,、 \t")
    else:
        remainder = raw[last_end:].strip().lstrip("，,、—- \t")
        nm2 = re.match(r"([^\s\d，,、]{1,10})\s+", remainder)
        if not nm2:
            raise ValueError("找唔到收件人名 — 請在電話前加上姓名")
        name = nm2.group(1).strip("，,、 \t")

    m = re.search(r"1[3-9]\d{9}|[2-9]\d{7}", raw)
    if not m:
        raise ValueError("找唔到電話號碼")
    phone = m.group(0)

    m = re.search(r"(?:1[3-9]\d{9}|[2-9]\d{7})\s*(.+)", raw, re.DOTALL)
    if not m:
        raise ValueError("找唔到地址")
    address = re.split(r"[。\n]", m.group(1).strip())[0].strip()

    return {
        "name": name, "phone": phone, "address": address,
        "items_pos": items_pos, "items_sf": items_sf,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Step 1 — 10 行輸入框
# ══════════════════════════════════════════════════════════════════════════════

def show_order_input() -> list:
    """彈出 10 行訂單輸入框，解析後返回 [order, ...] list。取消則退出程式。"""
    results = [None]

    root = tk.Tk()
    root.title("順丰寄件 — 輸入客人訂單")
    root.resizable(True, True)
    root.geometry("900x820")

    # ── 可滾動輸入框 ───────────────────────────────────────────────────────
    input_lf = tk.LabelFrame(
        root,
        text="WhatsApp 訂單 (每行一個客人，貼入後按解析)",
        padx=6, pady=6,
    )
    input_lf.pack(fill="both", padx=12, pady=(12, 4), expand=False)

    canvas_frame = tk.Canvas(input_lf, height=min(_NUM_ROWS * 28, 420))
    vsb = tk.Scrollbar(input_lf, orient="vertical", command=canvas_frame.yview)
    canvas_frame.configure(yscrollcommand=vsb.set)
    vsb.pack(side="right", fill="y")
    canvas_frame.pack(side="left", fill="both", expand=True)

    inner = tk.Frame(canvas_frame)
    canvas_frame.create_window((0, 0), window=inner, anchor="nw")
    inner.bind("<Configure>",
               lambda e: canvas_frame.configure(
                   scrollregion=canvas_frame.bbox("all")))
    canvas_frame.bind("<MouseWheel>",
                      lambda e: canvas_frame.yview_scroll(-1*(e.delta//120), "units"))

    row_vars   = []
    row_status = []

    for i in range(_NUM_ROWS):
        sv  = tk.StringVar()
        rst = tk.StringVar(value="")
        row_vars.append(sv)
        row_status.append(rst)

        row_frame = tk.Frame(inner)
        row_frame.pack(fill="x", pady=1)

        tk.Label(row_frame, text=f"{i+1:2d}.", width=3,
                 anchor="e", font=("Courier", 10)).pack(side="left")
        tk.Entry(row_frame, textvariable=sv,
                 font=("", 10), relief="solid", bd=1).pack(
            side="left", fill="x", expand=True, padx=(4, 4))
        tk.Label(row_frame, textvariable=rst,
                 width=10, anchor="w", font=("", 9)).pack(side="left")

    btn_row_top = tk.Frame(input_lf)
    btn_row_top.pack(fill="x", pady=(6, 0), side="bottom")
    tk.Button(btn_row_top, text="🗑 清除全部",
              command=lambda: [sv.set("") or rst.set("")
                               for sv, rst in zip(row_vars, row_status)],
              bg="#95a5a6", fg="white", padx=8, pady=3).pack(side="left")

    # ── Parsed results table ───────────────────────────────────────────────
    result_lf = tk.LabelFrame(root, text="已解析訂單", padx=6, pady=4)
    result_lf.pack(fill="both", expand=True, padx=12, pady=4)

    cols = ("#", "收件人", "電話", "貨品", "狀態")
    tree = ttk.Treeview(result_lf, columns=cols, show="headings", height=8)
    tree.heading("#",    text="#");    tree.column("#",    width=30, anchor="center")
    tree.heading("收件人", text="收件人"); tree.column("收件人", width=90)
    tree.heading("電話",  text="電話");  tree.column("電話",  width=120)
    tree.heading("貨品",  text="貨品");  tree.column("貨品",  width=380)
    tree.heading("狀態",  text="狀態");  tree.column("狀態",  width=100, anchor="center")
    tree.pack(fill="both", expand=True)
    tree.tag_configure("ok",    foreground="#27ae60")
    tree.tag_configure("error", foreground="#e74c3c")

    parsed_orders = []

    def on_parse():
        try:
            parsed_orders.clear()
            for item in tree.get_children():
                tree.delete(item)

            ok_count = 0
            for i, sv in enumerate(row_vars):
                raw = sv.get().strip()
                if not raw:
                    row_status[i].set("")
                    continue
                try:
                    order = _parse_order(raw)
                    items_str = ", ".join(
                        f"{it['name']}×{it['qty']}"
                        for it in order["items_sf"])
                    parsed_orders.append(order)
                    tree.insert("", "end",
                                values=(i+1, order["name"], order["phone"],
                                        items_str, "待寄出"),
                                tags=("ok",))
                    row_status[i].set("✅")
                    ok_count += 1
                except ValueError as e:
                    row_status[i].set("❌")
                    tree.insert("", "end",
                                values=(i+1, "—", "—", str(e)[:60], "解析失敗"),
                                tags=("error",))

            if ok_count:
                run_btn.config(state="normal")
                info_var.set(f"✅ 解析完成：{ok_count} 個訂單準備好，按「開始自動化」")
            else:
                run_btn.config(state="disabled")
                info_var.set("❌ 沒有成功解析的訂單")

        except Exception as e:
            messagebox.showerror("解析錯誤", f"內部錯誤：{e}")

    def on_run():
        if parsed_orders:
            results[0] = list(parsed_orders)
            root.destroy()

    def on_cancel():
        sys.exit(0)

    # ── Info label ─────────────────────────────────────────────────────────
    info_var = tk.StringVar(value="（貼入訂單後按「解析全部訂單」）")
    tk.Label(root, textvariable=info_var, anchor="w",
             fg="#555", font=("", 9)).pack(fill="x", padx=14)

    # ── Bottom buttons ─────────────────────────────────────────────────────
    btn_row = tk.Frame(root)
    btn_row.pack(pady=(4, 14))

    tk.Button(btn_row, text="🔍 解析全部訂單",
              command=on_parse,
              bg="#2980b9", fg="white",
              padx=14, pady=6, font=("", 11)).pack(side="left", padx=6)

    run_btn = tk.Button(btn_row, text="🚀 開始自動化",
                        command=on_run,
                        bg="#27ae60", fg="white",
                        padx=14, pady=6, font=("", 11, "bold"),
                        state="disabled")
    run_btn.pack(side="left", padx=6)

    tk.Button(btn_row, text="✖ 取消",
              command=on_cancel,
              bg="#e74c3c", fg="white",
              padx=14, pady=6, font=("", 11)).pack(side="left", padx=6)

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()
    return results[0]


# ══════════════════════════════════════════════════════════════════════════════
# 取得訂單資料
# ══════════════════════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════════════════════
# py6 函數：重新列印運單（行完 py5 自動接著行）
# ══════════════════════════════════════════════════════════════════════════════

def show_reprint_confirm_gui(session: list) -> bool:
    """py5 完成後彈出確認視窗，一鍵開始重新列印。"""
    result = [False]
    root = tk.Tk()
    root.title("重新列印順丰運單")
    root.geometry("680x320")

    tk.Label(root,
             text=f"py5 已完成 {len(session)} 張單，現在重新列印覆蓋舊 PDF",
             font=("", 12, "bold"), anchor="w").pack(fill="x", padx=12, pady=(12, 4))

    frame = tk.Frame(root)
    frame.pack(fill="both", expand=True, padx=12, pady=4)
    cols = ("#", "客人名", "順丰運單號", "運單 PDF 路徑")
    tree = ttk.Treeview(frame, columns=cols, show="headings", height=6)
    for c in cols:
        tree.heading(c, text=c)
    tree.column("#",          width=30,  anchor="center")
    tree.column("客人名",     width=90)
    tree.column("順丰運單號", width=160)
    tree.column("運單 PDF 路徑", width=340)
    tree.pack(fill="both", expand=True, side="left")
    ttk.Scrollbar(frame, orient="vertical", command=tree.yview).pack(side="right", fill="y")
    for i, d in enumerate(session):
        tree.insert("", "end", values=(i+1, d.get("customer",""), d.get("waybill",""), d.get("pdf_path","")))

    tk.Label(root, text="流程：修改訂單 → 直接保存 → 列印面單 → 覆蓋原 PDF",
             fg="#555", font=("", 9)).pack(anchor="w", padx=14)

    def on_confirm(): result[0] = True; root.destroy()
    def on_skip():    result[0] = False; root.destroy()

    btn_row = tk.Frame(root)
    btn_row.pack(pady=(8, 14))
    tk.Button(btn_row, text="開始重新列印", command=on_confirm,
              bg="#27ae60", fg="white", padx=16, pady=6, font=("", 11, "bold")).pack(side="left", padx=8)
    tk.Button(btn_row, text="跳過 / 稍後手動跑 py6", command=on_skip,
              bg="#95a5a6", fg="white", padx=16, pady=6, font=("", 10)).pack(side="left", padx=8)

    root.protocol("WM_DELETE_WINDOW", on_skip)
    root.mainloop()
    return result[0]


def _git_sync(message: str):
    """Commit tracking.xlsx and push to GitHub to update Streamlit."""
    _REPO = r"C:\Users\user\Desktop\順丰E順递"
    try:
        subprocess.run(["git", "-C", _REPO, "add", "data/tracking.xlsx"],
                       capture_output=True)
        subprocess.run(["git", "-C", _REPO, "commit", "-m", message],
                       capture_output=True)
        subprocess.run(["git", "-C", _REPO, "push", "origin", "main"],
                       capture_output=True)
        print(f"  ☁️  {message} — 已同步 Streamlit")
    except Exception as e:
        print(f"  ⚠️  Streamlit 同步失敗: {e}")


def mark_cancelled_in_excel(waybill: str):
    """Mark waybill as 已取消 in tracking.xlsx and push to Streamlit."""
    _ensure_excel()
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb[EXCEL_SHEET]
        col_wb = _HDR_COL["順丰運單號"]
        col_st = _HDR_COL["最新狀態"]
        found = False
        for row in ws.iter_rows(min_row=2):
            if str(row[col_wb - 1].value or "").strip() == waybill:
                row[col_st - 1].value = "已取消"
                found = True
                break
        if found:
            wb.save(EXCEL_PATH)
            print(f"  📊 Excel 已標記已取消: {waybill}")
            _git_sync(f"cancel: {waybill}")
        else:
            print(f"  ⚠️  Excel 找不到此運單: {waybill}")
    except Exception as e:
        print(f"  ⚠️  標記取消失敗: {e}")


def _is_waybill_cancelled(page) -> bool:
    """Check if the current SF page shows the order as cancelled."""
    try:
        body = page.inner_text("body")
        cancel_kw = ["已取消", "取消訂單", "訂單已取消", "訂單取消", "已撤銷"]
        return any(kw in body for kw in cancel_kw)
    except Exception:
        return False


def _dismiss_popups(page):
    try:
        page.evaluate("""() => {
            const dismiss = ['不用了','關閉','No thanks','Dismiss','取消','否'];
            for (const el of document.querySelectorAll('button')) {
                const t = (el.textContent || '').trim();
                if (dismiss.some(d => t.includes(d))) { el.click(); return; }
            }
        }""")
    except Exception:
        pass
    try: page.keyboard.press("Escape")
    except Exception: pass
    time.sleep(0.5)


def _go_to_waybill_detail(page, waybill: str):
    """列表頁 Ctrl+F 找運單號 → 點入詳情。"""
    page.goto(SF_LIST_URL, wait_until="domcontentloaded", timeout=30000)
    time.sleep(3)
    _dismiss_popups(page)

    page.keyboard.press("Control+f")
    time.sleep(0.8)
    page.keyboard.type(waybill, delay=80)
    time.sleep(1.0)
    page.keyboard.press("Escape")
    time.sleep(0.8)

    for attempt in range(8):
        try:
            page.get_by_text(waybill, exact=False).first.click(timeout=3000)
            print(f"     ✅ 點入運單（attempt {attempt+1}）")
            time.sleep(3)
            _dismiss_popups(page)
            return
        except Exception:
            pass
        time.sleep(1.5)

    print(f"     ⚠️  改用直接 URL")
    page.goto(f"{SF_DETAIL_BASE}/{waybill}", wait_until="domcontentloaded", timeout=30000)
    time.sleep(3)
    _dismiss_popups(page)


def reprint_one_waybill(ctx, entry: dict) -> bool:
    """
    SF 列表頁找運單 → 修改訂單 → 直接保存
    → 列印電子運單 → modal → 列印面單 → CDP 儲存 PDF 覆蓋舊檔
    """
    waybill  = str(entry.get("waybill",  "")).strip()
    customer = str(entry.get("customer", "")).strip()
    pdf_path = str(entry.get("pdf_path", "")).strip()

    print(f"\n{'='*55}")
    print(f"  重印: {waybill}  客人: {customer}")
    print(f"  覆蓋: {pdf_path}")
    print(f"{'='*55}")

    page = ctx.new_page()
    try:
        # 1. 列表頁找運單並點入
        print("  1. 從列表頁找運單")
        _go_to_waybill_detail(page, waybill)

        # ── 偵測訂單是否已取消 ────────────────────────────────────────────────
        if _is_waybill_cancelled(page):
            print(f"  ❌ 運單 {waybill} 已取消 — 自動更新 Streamlit 並跳過重印")
            page.close()
            mark_cancelled_in_excel(waybill)
            return False

        # 2. 點「修改訂單」
        print("  2. 點「修改訂單」")
        clicked = False
        for attempt in range(8):
            clicked = page.evaluate("""() => {
                for (const el of document.querySelectorAll('button,a,[role="button"],span,div')) {
                    if (el.offsetParent === null) continue;
                    const t = (el.textContent || '').trim();
                    if (t === '修改訂單' || t === '修改订单') { el.click(); return true; }
                }
                return false;
            }""")
            if clicked: print(f"     OK (attempt {attempt+1})"); break
            time.sleep(1.5)
        if not clicked: print("     找不到「修改訂單」，跳過"); return False
        time.sleep(3)

        # 3. 直接點「保存」
        print("  3. 直接點「保存」")
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(0.8)
        saved = False
        for attempt in range(8):
            saved = page.evaluate("""() => {
                const targets = ['保存','儲存'];
                for (const el of document.querySelectorAll('button,[role="button"]')) {
                    if (el.offsetParent === null) continue;
                    const t = (el.textContent || '').trim();
                    if (t === '取消' || t === '取消寄件') continue;
                    if (targets.some(c => t === c || t.endsWith(c))) { el.click(); return true; }
                }
                return false;
            }""")
            if saved: print(f"     OK (attempt {attempt+1})"); break
            time.sleep(1.5)
        if not saved: print("     找不到「保存」，跳過"); return False
        print("     等待成功頁面...")
        time.sleep(5)

        # 4. 點「列印電子運單」
        print("  4. 點「列印電子運單」")
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(0.5)
        clicked_print = False
        for attempt in range(6):
            clicked_print = page.evaluate("""() => {
                const labels = ['列印電子運單','打印電子運單'];
                for (const el of document.querySelectorAll('button,a,[role="button"],span')) {
                    if (el.offsetParent === null) continue;
                    const t = (el.textContent || '').trim();
                    if (labels.some(l => t === l || t.includes(l))) { el.click(); return true; }
                }
                return false;
            }""")
            if clicked_print: print(f"     OK (attempt {attempt+1})"); break
            time.sleep(1.5)
        if not clicked_print: print("     找不到「列印電子運單」，跳過"); return False
        print("     等待 modal...")
        time.sleep(3)

        # 5. 在 modal 點「列印面單」→ 新 tab → CDP 儲存 PDF
        print("  5. 點「列印面單」並儲存 PDF")
        try:
            with page.context.expect_page(timeout=20000) as new_pg_info:
                page.evaluate("""() => {
                    const labels = ['列印面單','打印面單'];
                    for (const el of document.querySelectorAll('button,a,[role="button"]')) {
                        if (el.offsetParent === null) continue;
                        const t = (el.textContent || '').trim();
                        if (labels.some(l => t === l || t.includes(l))) { el.click(); return true; }
                    }
                    return false;
                }""")
            print_page = new_pg_info.value
            print("     新 tab 已開啟，等待載入...")
            print_page.wait_for_load_state("domcontentloaded", timeout=15000)
            time.sleep(3)

            cdp = page.context.new_cdp_session(print_page)
            res = cdp.send("Page.printToPDF", {
                "printBackground": True, "preferCSSPageSize": True,
                "paperWidth": 8.27, "paperHeight": 11.69,
                "marginTop": 0, "marginBottom": 0, "marginLeft": 0, "marginRight": 0,
            })
            pdf_bytes = base64.b64decode(res["data"])
            os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
            with open(pdf_path, "wb") as f:
                f.write(pdf_bytes)
            cdp.detach()
            print_page.close()
            print(f"  ✅ PDF 已覆蓋：{pdf_path}")
        except Exception as pe:
            print(f"     PDF 儲存失敗：{pe}")

        return True

    except Exception as e:
        print(f"\n  {customer} ({waybill}) 失敗：{e}")
        import traceback; traceback.print_exc()
        return False
    finally:
        try: page.close()
        except Exception: pass


orders = show_order_input()   # list of order dicts

# ─── 關閉殘留 Chrome 進程 & 清理 lock files ──────────────────────────────────────

# 只 kill 用 ChromeAutomation profile 的 Chrome，唔影響其他 Chrome 視窗
subprocess.run(
    ["powershell", "-Command",
     "Get-WmiObject Win32_Process | Where-Object { $_.CommandLine -like '*ChromeAutomation*' } | ForEach-Object { $_.Terminate() }"],
    capture_output=True
)
time.sleep(1.5)

for lf in ["lockfile", "SingletonLock", "SingletonSocket", "SingletonCookie"]:
    try:
        os.remove(os.path.join(CHROME_PROFILE, lf))
    except Exception:
        pass

os.makedirs(LOGS_DIR, exist_ok=True)
today = date.today().strftime("%Y%m%d")


# ══════════════════════════════════════════════════════════════════════════════
# 工具函數
# ══════════════════════════════════════════════════════════════════════════════

def shot(page, label):
    p = os.path.join(LOGS_DIR, f"demo_{label}.png")
    page.screenshot(path=p, full_page=False)
    print(f"  📸 {p}")


def js_click_text(page, *texts, min_w=20, min_h=10):
    js_texts = str(list(texts))
    return page.evaluate(f"""() => {{
        const targets = {js_texts};
        for (const el of document.querySelectorAll('*')) {{
            if (el.offsetParent === null || el.children.length > 0) continue;
            const t = el.textContent.trim();
            if (!targets.includes(t)) continue;
            const r = el.getBoundingClientRect();
            if (r.width < {min_w} || r.height < {min_h}) continue;
            el.click();
            return {{ok: true, text: t}};
        }}
        return {{ok: false}};
    }}""")


def js_fill_by_label(page, label, value):
    info = page.evaluate(f"""() => {{
        const dialog = document.querySelector('[role="dialog"][data-state="open"]');
        const root = dialog || document;
        const targets = ['{label}', '{label}：', '{label}:'];
        const labels = [];
        for (const el of root.querySelectorAll('*')) {{
            if (el.offsetParent === null) continue;
            const t = el.textContent.trim();
            if (targets.includes(t)) {{ labels.push(el); continue; }}
            if (t.length < 12 && t.includes('{label}')) labels.push(el);
        }}
        if (!labels.length) return null;
        labels.sort((a,b) => a.children.length - b.children.length);
        for (const lbl of labels) {{
            let p = lbl.parentElement;
            for (let d = 0; d < 5; d++) {{
                if (!p) break;
                for (const inp of p.querySelectorAll('input:not([type=hidden]):not([type=radio]):not([type=checkbox])')) {{
                    if (inp.offsetParent === null) continue;
                    const r = inp.getBoundingClientRect();
                    if (r.width === 0) continue;
                    return {{x: r.left + r.width/2, y: r.top + r.height/2}};
                }}
                p = p.parentElement;
            }}
        }}
        return null;
    }}""")
    if not info:
        print(f"  ⚠️  找不到欄位: {label}")
        return False
    page.mouse.click(info["x"], info["y"])
    time.sleep(0.3)
    page.keyboard.press("Control+A")
    page.keyboard.press("Delete")
    page.keyboard.type(str(value), delay=30)
    time.sleep(0.3)
    return True


def smart_fill(page, text, which):
    page.locator("span:has-text('智慧填寫')").nth(which).click()
    time.sleep(0.8)
    ta = page.locator("textarea[class*='intelAddr'], textarea[placeholder*='陳先生']").first
    ta.wait_for(state="visible", timeout=5000)
    ta.click()
    ta.type(text, delay=20)
    time.sleep(0.5)
    page.evaluate("""() => {
        for (const el of document.querySelectorAll('*')) {
            if (el.childNodes.length===1 && el.firstChild.nodeType===3
                && el.firstChild.textContent.trim()==='識別') {
                el.click(); return;
            }
        }
    }""")
    time.sleep(2.5)


def download_pos_word(ctx, pos_order_no, save_dir, customer_name):
    """返 POS 銷售記錄，搵到訂單，下載 Word清單 + 收貨明細 + 清關PDF 到客人 folder。"""
    pos2 = ctx.new_page()
    try:
        pos2.goto(POS_URL, wait_until="domcontentloaded", timeout=20000)
        time.sleep(3)

        # 登入後台管理
        pos2.locator("button:has-text('后台管理')").first.click()
        time.sleep(0.8)
        pos2.locator("input[type='password']").first.fill(POS_PASS)
        pos2.keyboard.press("Enter")
        time.sleep(2)

        # 點「記錄」nav tab
        pos2.evaluate("""() => {
            for (const el of document.querySelectorAll('button')) {
                if (el.offsetParent === null) continue;
                const spans = [...el.querySelectorAll('span')];
                if (spans.some(s => s.textContent.trim() === '記錄')
                    || el.textContent.trim() === '記錄') {
                    el.click(); return;
                }
            }
        }""")
        time.sleep(2)
        print("  ✅ 進入銷售記錄")

        # 搜尋訂單號
        search = pos2.locator("input[placeholder*='搜尋單號']").first
        search.wait_for(state="visible", timeout=5000)
        search.click()
        # 先等 orders 從 Firebase 載入
        try:
            pos2.wait_for_function(
                "() => document.querySelectorAll('a[download]').length > 0",
                timeout=15000)
        except Exception:
            pass
        # 搜尋（逐字輸入確保 React filter 更新）
        search.click()
        search.triple_click()
        search.type(pos_order_no, delay=80)
        time.sleep(4)
        print(f"  [OK] 搜尋 {pos_order_no}")

        file_base = f"{customer_name}_{today}_{pos_order_no}"

        # ── 1. Word 清單 ──────────────────────────────────────────────────────
        word_path = os.path.join(save_dir, f"{file_base}_清單.docx")
        try:
            with pos2.expect_download(timeout=10000) as dl_info:
                pos2.evaluate("""() => {
                    for (const btn of document.querySelectorAll('button')) {
                        if (btn.offsetParent === null) continue;
                        const t = (btn.textContent || '').trim();
                        if (t.includes('Word清單') || t.includes('Word')) {
                            btn.click(); return;
                        }
                    }
                }""")
            dl = dl_info.value
            dl.save_as(word_path)
            print(f"  ✅ Word 清單已儲存: {word_path}")
        except Exception as dl_err:
            print(f"  ⚠️  Word 下載失敗: {dl_err}")

        # ── 收貨明細 / 清關：Ctrl+F 鍵盤模擬輔助函數 ────────────────────────
        def _ctrlf_find(page, order_no):
            """
            用鍵盤模擬 Ctrl+F 開瀏覽器搜尋欄，輸入訂單號讓瀏覽器
            scroll 到該訂單位置，等 3 秒讓 blob PDF 生成，再按 Escape 關閉。
            """
            print(f"    [Ctrl+F] 鍵盤搜尋 {order_no}...")
            page.keyboard.press("Control+f")
            time.sleep(0.8)
            page.keyboard.type(order_no, delay=60)
            time.sleep(3)          # ← 等 3 秒，讓頁面 scroll 到訂單 + blob 生成
            page.keyboard.press("Escape")
            time.sleep(0.8)
            print(f"    [Ctrl+F] 完成，開始嘗試 Download")

        # ── 2. 收貨明細 PDF ──────────────────────────────────────────────────
        # DevTools 顯示 download attribute = "明細_ORD-XXXXX.pdf"
        print("\n  ── 下載收貨明細 ──")
        _ctrlf_find(pos2, pos_order_no)   # Ctrl+F 先搵位 + 等 blob 就緒

        try:
            # 等 blob href 確認就緒（最多 15 秒）
            try:
                pos2.wait_for_function(f"""() => {{
                    for (const a of document.querySelectorAll('a[download]')) {{
                        const dl = a.getAttribute('download') || '';
                        if (dl.includes('{pos_order_no}') && dl.includes('明細')
                            && a.href && a.href.startsWith('blob:')) return true;
                    }}
                    return false;
                }}""", timeout=15000)
                print("  [OK] 收貨明細 blob 已就緒")
            except Exception:
                print("  [WARN] 等待收貨明細 blob 超時，嘗試直接點擊")

            with pos2.expect_download(timeout=15000) as dl_info:
                pos2.evaluate(f"""() => {{
                    for (const a of document.querySelectorAll('a[download]')) {{
                        if (a.offsetParent === null) continue;
                        const dl = a.getAttribute('download') || '';
                        if (dl.includes('{pos_order_no}') && dl.includes('明細')) {{
                            a.click(); return true;
                        }}
                    }}
                }}""")
            dl = dl_info.value
            ext = os.path.splitext(dl.suggested_filename)[1] or ".pdf"
            receipt_path = os.path.join(save_dir, f"{file_base}_收貨明細{ext}")
            dl.save_as(receipt_path)
            print(f"  ✅ 收貨明細已儲存: {receipt_path}")
        except Exception as dl_err:
            print(f"  ⚠️  收貨明細下載失敗: {dl_err}")

        # ── 3. 清關 PDF ──────────────────────────────────────────────────────
        print("\n  ── 下載清關PDF ──")
        _ctrlf_find(pos2, pos_order_no)   # 再做一次 Ctrl+F 確保 blob 就緒

        try:
            try:
                pos2.wait_for_function(f"""() => {{
                    for (const a of document.querySelectorAll('a[download]')) {{
                        const dl = a.getAttribute('download') || '';
                        if (dl.includes('{pos_order_no}') && dl.includes('清關')
                            && a.href && a.href.startsWith('blob:')) return true;
                    }}
                    return false;
                }}""", timeout=15000)
                print("  [OK] 清關 blob 已就緒")
            except Exception:
                print("  [WARN] 等待清關 blob 超時，嘗試直接點擊")

            with pos2.expect_download(timeout=15000) as dl_info:
                pos2.evaluate(f"""() => {{
                    for (const a of document.querySelectorAll('a[download]')) {{
                        if (a.offsetParent === null) continue;
                        const dl = a.getAttribute('download') || '';
                        if (dl.includes('{pos_order_no}') && dl.includes('清關')) {{
                            a.click(); return true;
                        }}
                    }}
                }}""")
            dl = dl_info.value
            ext = os.path.splitext(dl.suggested_filename)[1] or ".pdf"
            customs_path = os.path.join(save_dir, f"{file_base}_清關{ext}")
            dl.save_as(customs_path)
            print(f"  ✅ 清關PDF已儲存: {customs_path}")
        except Exception as dl_err:
            print(f"  ⚠️  清關PDF下載失敗: {dl_err}")

    except Exception as e:
        print(f"  ⚠️  POS 文件下載失敗: {e}")
    finally:
        pos2.close()


def wait_dialog_closed(page, timeout=6):
    for _ in range(timeout * 2):
        gone = page.evaluate("""() =>
            document.querySelectorAll('[role="dialog"][data-state="open"]').length === 0
        """)
        if gone:
            return True
        time.sleep(0.5)
    return False


# ══════════════════════════════════════════════════════════════════════════════
# Excel 追蹤表工具
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_HEADERS = [
    "日期", "客人名", "POS訂單號", "順丰運單號",
    "收件人", "收件電話", "收件地址",
    "貨品摘要", "件數", "VIP總額(HKD)",
    "付款方式", "運費(HKD)", "最新狀態",
    "狀態更新時間", "異常標記", "小票檔案路徑", "備註", "稅金(HKD)",
]
EXCEL_SHEET = "追蹤表"

_HDR_COL = {h: i+1 for i, h in enumerate(EXCEL_HEADERS)}   # header → col index

def _ensure_excel():
    """建立 Excel 檔案（若不存在），設定標題行格式。"""
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET

    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx, hdr in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border

    # 欄寬
    col_widths = {
        "日期": 12, "客人名": 12, "POS訂單號": 14, "順丰運單號": 20,
        "收件人": 12, "收件電話": 14, "收件地址": 36,
        "貨品摘要": 34, "件數": 7, "VIP總額(HKD)": 13,
        "付款方式": 12, "運費(HKD)": 12, "最新狀態": 14,
        "狀態更新時間": 20, "異常標記": 10, "小票檔案路徑": 40,
        "備註": 16, "稅金(HKD)": 12,
    }
    for hdr, width in col_widths.items():
        ws.column_dimensions[get_column_letter(_HDR_COL[hdr])].width = width

    ws.freeze_panes = "A2"          # 鎖定標題行
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"  📊 Excel 表已建立: {EXCEL_PATH}")


def append_order_to_excel(order: dict, waybill: str, pdf_path: str, pos_order_no: str):
    """
    把一個訂單寫入 Excel 追蹤表（新增一行）。
    order keys: name, phone, items_sf
    """
    _ensure_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[EXCEL_SHEET]

    items_str  = " / ".join(
        f"{it['name']}×{it['qty']}" for it in order.get("items_sf", []))
    total_amt  = sum(
        it.get("unit_price", 0) * it.get("qty", 1)
        for it in order.get("items_sf", []))
    total_qty  = sum(it.get("qty", 1) for it in order.get("items_sf", []))

    row_data = {
        "日期":          today,
        "客人名":        order.get("name", ""),
        "POS訂單號":     pos_order_no,
        "順丰運單號":    waybill,
        "收件人":        order.get("name", ""),
        "收件電話":      order.get("phone", ""),
        "收件地址":      order.get("address", ""),
        "貨品摘要":      items_str,
        "件數":          total_qty,
        "VIP總額(HKD)":  round(total_amt, 1),
        "付款方式":      "月結",
        "最新狀態":      "待更新",
        "小票檔案路徑":  pdf_path,
    }

    # 插入第 2 行（header 下面），最新記錄永遠在最頂
    ws.insert_rows(2)
    next_row = 2
    data_fill   = PatternFill("solid", fgColor="EBF3FB")
    alt_fill    = PatternFill("solid", fgColor="FFFFFF")
    fill        = data_fill
    thin_side   = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)

    for hdr, col in _HDR_COL.items():
        cell = ws.cell(row=next_row, column=col, value=row_data.get(hdr, ""))
        cell.fill   = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    wb.save(EXCEL_PATH)
    print(f"  📊 Excel 已記錄: Row {next_row}  {order['name']}  {waybill}")


# ══════════════════════════════════════════════════════════════════════════════
# 主流程：逐個訂單執行
# ══════════════════════════════════════════════════════════════════════════════

success_count = 0
failed_count  = 0
completed_orders = []   # [(customer, waybill, pdf_path)]

with sync_playwright() as pw:
    ctx = pw.chromium.launch_persistent_context(
        CHROME_PROFILE, channel="chrome", headless=False,
        args=["--disable-blink-features=AutomationControlled"],
        slow_mo=150, viewport={"width": 1280, "height": 900},
    )

    for order_idx, order in enumerate(orders):
        DEMO_CUSTOMER  = order["name"]
        DEMO_PHONE     = order["phone"]
        DEMO_ADDRESS   = order["address"]
        DEMO_POS_ITEMS = order["items_pos"]
        DEMO_SF_ITEMS  = order["items_sf"]

        print(f"\n{'='*60}")
        print(f"  訂單 {order_idx+1}/{len(orders)}: {DEMO_CUSTOMER}")
        print(f"{'='*60}")

        try:
            # ════════════════════════════════════════════════════════════════
            # 階段一：POS 打小票 (Steps 1-5)
            # ════════════════════════════════════════════════════════════════

            print("\n▶ Step 2: 登入 POS 後台")
            pos_page = ctx.new_page()
            pos_page.goto(POS_URL, wait_until="domcontentloaded", timeout=20000)
            time.sleep(3)

            pos_page.locator("button:has-text('后台管理')").first.click()
            time.sleep(0.8)
            pos_page.locator("input[type='password']").first.fill(POS_PASS)
            pos_page.keyboard.press("Enter")
            time.sleep(1.5)
            print("  ✅ 後台管理已登入")

            print("\n▶ Step 3: 啟動 VIP 價")
            pos_page.locator("button:has-text('VIP價')").first.click()
            time.sleep(0.8)
            pos_page.locator("input[type='password']").first.fill(VIP_PASS)
            pos_page.keyboard.press("Enter")
            time.sleep(1.5)
            print("  ✅ VIP 價已啟動")

            print(f"\n▶ Step 3: 加入 {len(DEMO_POS_ITEMS)} 種貨品")
            # 等 POS 貨品從 Firebase 載入完畢（至少第一個 SKU 出現在某個 button 裏）
            first_sku = DEMO_POS_ITEMS[0]["sku"]
            try:
                pos_page.wait_for_function(
                    f"() => {{ for(const b of document.querySelectorAll('button')) {{ if(b.textContent.includes('{first_sku}')) return true; }} return false; }}",
                    timeout=20000)
                print("  POS 貨品已載入")
            except Exception:
                print("  等待 POS 貨品載入超時，繼續嘗試")
            for item in DEMO_POS_ITEMS:
                try:
                    btn = pos_page.locator(f"button:has-text('{item['sku']}')").first
                    btn.wait_for(state="visible", timeout=15000)
                    for _ in range(item["qty"]):
                        btn.click()
                        time.sleep(0.25)
                    print(f"  SKU {item['sku']} x {item['qty']}")
                except Exception as e:
                    print(f"  [跳過] SKU {item['sku']} 未找到: {e}")

            shot(pos_page, f"{order_idx+1:02d}_01_pos_cart")

            print("\n▶ Step 4: 結帳 (VIP 價)")
            pos_page.locator("button:has-text('結帳')").first.click()
            time.sleep(1.5)

            # 現金係 default，直接按確認出小票
            pos_page.locator("button:has-text('確認，出小票')").first.click()
            time.sleep(2.5)
            shot(pos_page, f"{order_idx+1:02d}_02_pos_receipt")

            body_text = pos_page.inner_text("body")
            m = re.search(r"ORD-\d+", body_text)
            pos_order_no = m.group(0) if m else f"ORD-{today}"
            print(f"  ✅ POS 訂單號: {pos_order_no}")

            print("\n▶ Step 5: 儲存小票到客人 Folder")
            order_folder_name = f"{DEMO_CUSTOMER}_{today}_{pos_order_no}"
            save_dir  = os.path.join(ORDERS_DIR, order_folder_name)  # git-tracked
            os.makedirs(save_dir, exist_ok=True)
            file_base = f"{DEMO_CUSTOMER}_{today}_{pos_order_no}"
            pdf_path  = os.path.join(save_dir, file_base + ".pdf")
            png_path  = os.path.join(save_dir, file_base + ".png")
            pdf_rel   = f"data/orders/{order_folder_name}/{file_base}.pdf"  # relative path for Excel + cloud

            # 按藍色 DOWNLOAD 按鈕，Playwright 攔截下載並儲存到客人 folder
            try:
                with pos_page.expect_download(timeout=15000) as dl_info:
                    pos_page.evaluate("""() => {
                        const cands = ['DOWNLOAD','Download','download','下載','下載小票','PDF','列印','Print'];
                        for (const t of cands) {
                            for (const el of document.querySelectorAll('button, a, [role="button"]')) {
                                if (el.offsetParent === null) continue;
                                const txt = el.textContent.trim();
                                if (txt === t || txt.includes(t)) { el.click(); return; }
                            }
                        }
                    }""")
                dl = dl_info.value
                dl.save_as(pdf_path)
                print(f"  ✅ 小票已儲存: {pdf_path}")
            except Exception as dl_err:
                print(f"  ⚠️  下載攔截失敗，改截圖備份: {dl_err}")
                pos_page.screenshot(path=png_path, full_page=False)
                pdf_path = png_path
                pdf_rel  = f"data/orders/{order_folder_name}/{file_base}.png"
                print(f"  ✅ 截圖備份: {png_path}")

            try:
                done_btn = pos_page.locator("button:has-text('完成')").first
                if done_btn.is_visible(timeout=2000):
                    done_btn.click()
                    time.sleep(0.8)
            except Exception:
                pass
            pos_page.close()

            # Step 6 (人手包裝) — 已跳過，直接進入順豐落單

            # ════════════════════════════════════════════════════════════════
            # 階段三：填順丰快遞單 (Steps 7-13)
            # ════════════════════════════════════════════════════════════════

            print("\n▶ Step 7: 開啟順豐寄件頁面")
            sf_page = ctx.new_page()
            sf_page.goto(SF_URL, wait_until="domcontentloaded", timeout=20000)
            time.sleep(3)

            print("\n▶ Step 8: 寄件人 → 自寄")
            smart_fill(sf_page, SENDER_TEXT, 0)

            sf_page.evaluate("""() => {
                const inputs = document.querySelectorAll("input[name='contactName']");
                if (inputs[0]) {
                    const setter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype,'value').set;
                    setter.call(inputs[0],'潘正儀');
                    inputs[0].dispatchEvent(new Event('input',{bubbles:true}));
                    inputs[0].dispatchEvent(new Event('change',{bubbles:true}));
                }
            }""")
            time.sleep(0.5)

            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    if (el.textContent.trim()==='自寄' && el.offsetParent!==null) {
                        const r = el.getBoundingClientRect();
                        if (r.width>40 && r.height>20) { el.click(); return; }
                    }
                }
            }""")
            time.sleep(1.5)
            print("  ✅ 自寄已選")

            print("\n▶ Step 9: 收件人智慧填寫")
            recip_text = f"{DEMO_CUSTOMER} {DEMO_PHONE} {DEMO_ADDRESS}"
            smart_fill(sf_page, recip_text, 1)
            shot(sf_page, f"{order_idx+1:02d}_03_sf_recipient")
            print("  ✅ 收件人已填")

            print(f"\n▶ Step 10: 填 {len(DEMO_SF_ITEMS)} 件物品信息")

            for idx, item in enumerate(DEMO_SF_ITEMS):
                print(f"\n  ── 物品 {idx+1}: {item['name']} ──")

                info = sf_page.evaluate("""() => {
                    const candidates = ['+新增物品','新增物品','+ 新增物品'];
                    for (const el of document.querySelectorAll('*')) {
                        const t = el.textContent.trim();
                        if (!candidates.includes(t)) continue;
                        if (el.offsetParent===null || el.children.length>0) continue;
                        const r = el.getBoundingClientRect();
                        if (r.width<20||r.height<10) continue;
                        return {absY: r.top+window.scrollY, h:r.height};
                    }
                    return null;
                }""")
                if info:
                    sf_page.evaluate(f"window.scrollTo(0,{info['absY']}-450+{info['h']}/2)")
                    time.sleep(0.8)
                click_info = sf_page.evaluate("""() => {
                    const candidates = ['+新增物品','新增物品','+ 新增物品'];
                    for (const el of document.querySelectorAll('*')) {
                        const t = el.textContent.trim();
                        if (!candidates.includes(t)) continue;
                        if (el.offsetParent===null || el.children.length>0) continue;
                        const r = el.getBoundingClientRect();
                        if (r.width<20||r.height<10) continue;
                        return {x:r.left+r.width/2, y:r.top+r.height/2};
                    }
                    return null;
                }""")
                if not click_info:
                    print("  ❌ 搵唔到 +新增物品")
                    break
                sf_page.mouse.click(click_info["x"], click_info["y"])
                time.sleep(2)
                # v4: 等 dialog 入面 input 真係 visible，確保 dialog 完全 render 完
                try:
                    sf_page.locator('[role="dialog"][data-state="open"] input').first.wait_for(state="visible", timeout=6000)
                except Exception:
                    pass
                time.sleep(0.5)

                sf_page.evaluate("""() => {
                    for (const el of document.querySelectorAll('*')) {
                        if (el.textContent.trim()==='物品' && el.offsetParent!==null
                            && el.children.length===0) {
                            const r = el.getBoundingClientRect();
                            if (r.width>10&&r.height>5) { el.click(); return; }
                        }
                    }
                }""")
                time.sleep(1.5)

                js_fill_by_label(sf_page, "物品名稱", item["name"])
                sf_page.keyboard.press("Tab")
                time.sleep(4)
                print("  ✅ 物品名稱")

                js_fill_by_label(sf_page, "品牌", item["brand"]); time.sleep(2)
                print("  ✅ 品牌")
                # 材質欄位：順豐可能叫「材質」或「物品材質」
                mat_filled = False
                for mat_label in ["材質", "物品材質"]:
                    if js_fill_by_label(sf_page, mat_label, item["material"]):
                        mat_filled = True
                        break
                if not mat_filled:
                    print("  ⚠️  材質欄位搵唔到，跳過")
                time.sleep(2)
                print("  ✅ 材質")
                # 規格欄位：順豐可能叫「規格型號」、「規格」或「型號」
                spec_filled = False
                for spec_label in ["規格型號", "規格", "型號"]:
                    if js_fill_by_label(sf_page, spec_label, item["spec"]):
                        spec_filled = True
                        break
                if not spec_filled:
                    print("  ⚠️  規格欄位搵唔到，跳過")
                time.sleep(2)
                print("  ✅ 規格")

                # 原產地 → 台灣
                # Dialog 裡有3個 combobox（港幣/個/原產地），最後一個係原產地
                origin_trigger = sf_page.evaluate("""() => {
                    const dialog = document.querySelector('[role="dialog"][data-state="open"]');
                    if (!dialog) return null;
                    const btns = [...dialog.querySelectorAll('button[role="combobox"]')]
                        .filter(el => el.offsetParent !== null);
                    if (btns.length === 0) return null;
                    const btn = btns[btns.length - 1];  // 最後一個 = 原產地
                    const r = btn.getBoundingClientRect();
                    return {x: r.left + r.width / 2, y: r.top + r.height / 2};
                }""")
                if origin_trigger:
                    sf_page.mouse.click(origin_trigger["x"], origin_trigger["y"])
                    time.sleep(1.0)
                    taiwan_opt = sf_page.evaluate("""() => {
                        const opts = [...document.querySelectorAll('[role="option"]')]
                            .filter(el => el.offsetParent !== null);
                        for (const opt of opts) {
                            const t = (opt.textContent || '').trim();
                            if (t.includes('台灣') || t.includes('臺灣')) {
                                const r = opt.getBoundingClientRect();
                                return {x: r.left + r.width / 2, y: r.top + r.height / 2, text: t};
                            }
                        }
                        if (opts.length >= 4) {
                            const r = opts[3].getBoundingClientRect();
                            return {x: r.left + r.width / 2, y: r.top + r.height / 2,
                                    text: opts[3].textContent.trim()};
                        }
                        return null;
                    }""")
                    if taiwan_opt:
                        sf_page.mouse.click(taiwan_opt["x"], taiwan_opt["y"])
                        print(f"  ✅ 原產地 → {taiwan_opt['text']}")
                    else:
                        print("  ⚠️  找唔到台灣選項，跳過原產地")
                    time.sleep(0.8)
                else:
                    print("  ⚠️  原產地下拉選單搵唔到，跳過")

                js_fill_by_label(sf_page, "物品單價", item["unit_price"]); time.sleep(0.5)
                print("  ✅ 單價")
                js_fill_by_label(sf_page, "物品數量", item["qty"]); time.sleep(0.5)
                print("  ✅ 數量")

                sf_page.evaluate("""() => {
                    for (const el of document.querySelectorAll('[class*="package-declaration_confirm"]')) {
                        if (el.offsetParent===null) continue;
                        if (el.textContent.trim()!=='確認') continue;
                        const r = el.getBoundingClientRect();
                        if (r.width<30||r.height<15) continue;
                        el.click(); return;
                    }
                }""")
                time.sleep(2)
                wait_dialog_closed(sf_page)
                print("  ✅ 已確認")
                # v4: scroll 返頂部，唔理預估總重量欄位
                sf_page.evaluate("window.scrollTo(0, 0)")
                time.sleep(2)

            # v4: 最後一件物品確認後，等系統 recalculate 申報總值先落單
            print("  ⏳ 等系統更新申報總值...")
            time.sleep(9)

            print("\n▶ Step 11: 付款方式 → 寄付月結")
            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    if (el.offsetParent===null || el.children.length>0) continue;
                    const t = el.textContent.trim();
                    if (t!=='月結' && t!=='寄付月結') continue;
                    const r = el.getBoundingClientRect();
                    if (r.width<20||r.height<15) continue;
                    el.click(); return;
                }
            }""")
            time.sleep(1.5)
            print("  ✅ 寄付月結已選")

            print(f"\n▶ Step 12: 填月結卡號 {MONTHLY_ACCOUNT}")
            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    if (el.textContent.trim()==='付款方式' && el.offsetParent!==null)
                        { el.scrollIntoView({block:'center'}); return; }
                }
            }""")
            time.sleep(0.8)
            acct_info = sf_page.evaluate("""() => {
                for (const inp of document.querySelectorAll('input')) {
                    if (inp.offsetParent===null) continue;
                    const ph = inp.placeholder||'';
                    const r = inp.getBoundingClientRect();
                    if (r.width===0||r.height===0) continue;
                    if (ph.includes('月結')||ph.includes('卡號'))
                        return {x:r.left+r.width/2, y:r.top+r.height/2};
                }
                return null;
            }""")
            if acct_info:
                sf_page.mouse.click(acct_info["x"], acct_info["y"])
                time.sleep(0.3)
                sf_page.mouse.click(acct_info["x"], acct_info["y"])
                time.sleep(0.3)
                sf_page.keyboard.press("Control+A")
                sf_page.keyboard.press("Delete")
                sf_page.keyboard.type(MONTHLY_ACCOUNT, delay=50)
                time.sleep(0.8)
                sf_page.evaluate(f"""() => {{
                    for (const el of document.querySelectorAll('*')) {{
                        if (el.offsetParent===null||el.children.length>0) continue;
                        const t = el.textContent.trim();
                        if (t.includes('{MONTHLY_ACCOUNT}')&&t.length<50) {{
                            const r=el.getBoundingClientRect();
                            if (r.width<30) return;
                            el.click(); return;
                        }}
                    }}
                }}""")
                sf_page.keyboard.press("Tab")
                time.sleep(0.5)
                print("  ✅ 月結卡號已填")
            else:
                print("  ⚠️  月結卡號欄位未找到")

            print("\n▶ Step 13a: 閱讀並同意")
            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('[class*="agreedCheckbox"]')) {
                    if (el.offsetParent!==null)
                        { el.scrollIntoView({block:'center'}); return; }
                }
            }""")
            time.sleep(0.8)
            sf_page.evaluate("""() => {
                for (const el of document.querySelectorAll('[class*="agreedCheckbox"]')) {
                    if (el.offsetParent===null) continue;
                    const r = el.getBoundingClientRect();
                    if (r.width<30) continue;
                    el.click(); return;
                }
                for (const el of document.querySelectorAll('[class*="checkbox_checkbox"]')) {
                    if (el.offsetParent===null) continue;
                    const r = el.getBoundingClientRect();
                    if (r.width<30) continue;
                    const txt = (el.parentElement||el).textContent||'';
                    if (txt.includes('閱讀')&&txt.includes('同意'))
                        { el.click(); return; }
                }
            }""")
            time.sleep(2)
            print("  ✅ Checkbox 已勾")

            sf_page.evaluate("""() => {
                const cands=['同意本條款,下次不再提示','同意本條款，下次不再提示',
                             '同意本條款','同意並繼續','同意'];
                for (const t of cands) {
                    for (const el of document.querySelectorAll('*')) {
                        if (el.offsetParent===null) continue;
                        if (el.textContent.trim()!==t) continue;
                        if (el.children.length>0) continue;
                        const r=el.getBoundingClientRect();
                        if (r.width<50||r.height<20) continue;
                        el.click(); return;
                    }
                }
            }""")
            time.sleep(1.5)

            print("\n▶ Step 13: 下單")
            sf_page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(0.5)

            waybill = "未擷取"
            submit_result = {}
            for attempt in range(8):
                submit_result = sf_page.evaluate("""() => {
                    for (const el of document.querySelectorAll('[class*="submitBtn"]')) {
                        if (el.offsetParent===null) continue;
                        if (!el.textContent.trim().includes('下單')) continue;
                        const cls=(el.className||'').toString();
                        if (cls.includes('disabled')) return {ok:false,reason:'disabled'};
                        const r=el.getBoundingClientRect();
                        if (r.width<40) continue;
                        el.click();
                        return {ok:true};
                    }
                    for (const el of document.querySelectorAll('[role="button"]')) {
                        if (el.offsetParent===null) continue;
                        if (el.textContent.trim()!=='下單') continue;
                        const cls=(el.className||'').toString();
                        if (cls.includes('disabled')) return {ok:false,reason:'role_disabled'};
                        el.click();
                        return {ok:true};
                    }
                    return {ok:false,reason:'not_found'};
                }""")
                if submit_result.get("ok"):
                    print(f"  ✅ 下單成功 (attempt {attempt+1})")
                    # v4: 等系統 load 確認頁，確保申報值已正確記錄
                    print("  ⏳ 等系統處理下單...")
                    time.sleep(5)
                    break
                time.sleep(2)
            else:
                print(f"  ❌ 下單失敗: {submit_result}")

            time.sleep(3)
            shot(sf_page, f"{order_idx+1:02d}_04_sf_submitted")

            # ── 跳過報關頁面（如彈出） ────────────────────────────────────────
            try:
                skip_customs = sf_page.evaluate("""() => {
                    const skipTexts = ['跳過','稍後填寫','暫不填寫','跳過報關'];
                    for (const t of skipTexts) {
                        for (const el of document.querySelectorAll('button, a, [role="button"]')) {
                            if (el.offsetParent===null) continue;
                            const txt = el.textContent.trim();
                            if (txt === t || txt.includes(t)) {
                                el.click(); return true;
                            }
                        }
                    }
                    return false;
                }""")
                if skip_customs:
                    print("  ✅ 報關頁已跳過")
                    time.sleep(1.5)
            except Exception:
                pass

            for _ in range(5):
                content = sf_page.content()
                m = re.search(r"SF\d{10,}", content)
                if m:
                    waybill = m.group(0)
                    break
                time.sleep(1)

            print(f"\n  ✅ {DEMO_CUSTOMER} 完成！運單號: {waybill}")
            print(f"  📄 小票: {pdf_path}")

            # ── 列印電子運單 → 儲存 PDF ──────────────────────────────────────
            try:
                print("\n▶ Step 14: 列印電子運單")
                waybill_pdf_name = f"{DEMO_CUSTOMER}_{today}_{pos_order_no}_{waybill}_運單.pdf"
                waybill_pdf_path = os.path.join(save_dir, waybill_pdf_name)

                # 直接導航到運單詳情頁
                print(f"  → 前往運單詳情頁: {SF_DETAIL_BASE}/{waybill}")
                sf_page.goto(f"{SF_DETAIL_BASE}/{waybill}",
                             wait_until="domcontentloaded", timeout=30000)
                time.sleep(3)
                _dismiss_popups(sf_page)

                # 必須先「修改訂單」→「保存」，「列印電子運單」才會出現
                print("  → 點「修改訂單」")
                for attempt in range(8):
                    clicked_mod = sf_page.evaluate("""() => {
                        for (const el of document.querySelectorAll('button,a,[role="button"],span,div')) {
                            if (el.offsetParent === null) continue;
                            const t = (el.textContent || '').trim();
                            if (t === '修改訂單' || t === '修改订单') { el.click(); return true; }
                        }
                        return false;
                    }""")
                    if clicked_mod:
                        print(f"     ✅ 已點修改訂單 (attempt {attempt+1})")
                        break
                    time.sleep(1.5)
                else:
                    raise Exception("找不到「修改訂單」按鈕")
                time.sleep(3)

                print("  → 點「保存」")
                sf_page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(0.8)
                for attempt in range(8):
                    saved_mod = sf_page.evaluate("""() => {
                        const targets = ['保存','儲存'];
                        for (const el of document.querySelectorAll('button,[role="button"]')) {
                            if (el.offsetParent === null) continue;
                            const t = (el.textContent || '').trim();
                            if (t === '取消' || t === '取消寄件') continue;
                            if (targets.some(c => t === c || t.endsWith(c))) { el.click(); return true; }
                        }
                        return false;
                    }""")
                    if saved_mod:
                        print(f"     ✅ 已點保存 (attempt {attempt+1})")
                        break
                    time.sleep(1.5)
                else:
                    raise Exception("找不到「保存」按鈕")
                print("     等待保存完成...")
                time.sleep(5)

                # 點「列印電子運單」（保存後才出現）
                print("  → 點「列印電子運單」")
                sf_page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(0.5)
                clicked_print = False
                for attempt in range(6):
                    clicked_print = sf_page.evaluate("""() => {
                        const labels = ['列印電子運單', '打印電子運單'];
                        for (const el of document.querySelectorAll('button, a, [role="button"], span')) {
                            if (el.offsetParent === null) continue;
                            const t = (el.textContent || '').trim();
                            if (labels.some(l => t === l || t.includes(l))) {
                                el.click(); return true;
                            }
                        }
                        return false;
                    }""")
                    if clicked_print:
                        print(f"     ✅ 已點列印電子運單 (attempt {attempt+1})")
                        break
                    time.sleep(1.5)

                if not clicked_print:
                    print("  ⚠️  找不到列印電子運單按鈕，跳過")
                else:
                    time.sleep(2.5)
                    with ctx.expect_page() as new_page_info:
                        sf_page.evaluate("""() => {
                            const labels = ['列印面單', '列印頁面', '打印面單', '打印頁面'];
                            for (const el of document.querySelectorAll('button, a, [role="button"]')) {
                                if (el.offsetParent === null) continue;
                                const t = (el.textContent || '').trim();
                                if (labels.some(l => t === l || t.includes(l))) {
                                    el.click(); return true;
                                }
                            }
                            return false;
                        }""")

                    print_page = new_page_info.value
                    print_page.wait_for_load_state("networkidle", timeout=10000)
                    time.sleep(1.5)

                    # 用 CDP 儲存列印頁面為 PDF
                    import base64 as _b64
                    cdp = ctx.new_cdp_session(print_page)
                    result = cdp.send("Page.printToPDF", {
                        "printBackground": True,
                        "paperWidth":  8.27,
                        "paperHeight": 11.69,
                        "marginTop":    0.2,
                        "marginBottom": 0.2,
                        "marginLeft":   0.2,
                        "marginRight":  0.2,
                    })
                    pdf_data = _b64.b64decode(result["data"])
                    with open(waybill_pdf_path, "wb") as f:
                        f.write(pdf_data)
                    cdp.detach()
                    print_page.close()
                    print(f"  ✅ 電子運單已儲存: {waybill_pdf_path}")

            except Exception as e:
                print(f"  ⚠️  列印電子運單失敗: {e}")

            # ── 寫入 Excel 追蹤表 ────────────────────────────────────────────
            append_order_to_excel(order, waybill, pdf_rel, pos_order_no)

            sf_page.close()

            # ── Step 15: 返 POS 銷售記錄下載 Word 清單 ──────────────────────
            print(f"\n▶ Step 15: 下載 POS 銷售記錄 Word 清單 ({pos_order_no})")
            download_pos_word(ctx, pos_order_no, save_dir, DEMO_CUSTOMER)

            # ── 同步全部檔案到 GitHub（Word 下載完先 push，確保三個齊全）────────
            try:
                _REPO = r"C:\Users\user\Desktop\順丰E順递"
                subprocess.run(
                    ["git", "-C", _REPO, "add", "data/tracking.xlsx",
                     f"data/orders/{order_folder_name}"],
                    capture_output=True, check=True)
                subprocess.run(
                    ["git", "-C", _REPO, "commit", "-m",
                     f"order: {DEMO_CUSTOMER} {pos_order_no} {waybill}"],
                    capture_output=True, check=True)
                subprocess.run(
                    ["git", "-C", _REPO, "push", "origin", "main"],
                    capture_output=True, check=True)
                print("  ☁️  小票 + 運單 + Word + 追蹤表 已同步到雲端 Streamlit")
            except Exception as _ge:
                print(f"  ⚠️  雲端同步失敗（唔影響本地）: {_ge}")

            success_count += 1
            completed_orders.append((DEMO_CUSTOMER, waybill, pdf_path))

        except Exception as e:
            print(f"\n  ❌ {DEMO_CUSTOMER} 失敗: {e}")
            failed_count += 1
            try:
                sf_page.close()
            except Exception:
                pass

    # ── 儲存本次 session 給 py6 使用 ────────────────────────────────────────
    _SESSION_FILE = r"C:\Users\user\Desktop\順丰E順递\data\last_session.json"
    _session = []
    for c, w, p in completed_orders:
        # p 係小票路徑（{customer}_{date}_{pos_no}.pdf）
        # 運單 PDF 路徑係 {customer}_{date}_{pos_no}_{waybill}_運單.pdf
        _receipt_dir  = os.path.dirname(p)
        _receipt_base = os.path.basename(p).replace(".pdf", "")
        _waybill_pdf  = os.path.join(_receipt_dir, f"{_receipt_base}_{w}_運單.pdf")
        _session.append({"customer": c, "waybill": w, "pdf_path": _waybill_pdf})
    try:
        os.makedirs(os.path.dirname(_SESSION_FILE), exist_ok=True)
        with open(_SESSION_FILE, "w", encoding="utf-8") as _sf:
            json.dump(_session, _sf, ensure_ascii=False, indent=2)
        print(f"\n  💾 Session 已儲存 → 可以跑 py6：{_SESSION_FILE}")
    except Exception as _se:
        print(f"\n  ⚠️  Session 儲存失敗：{_se}")

    # ── 完成摘要 ────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  批量完成：✅ {success_count} 成功  ❌ {failed_count} 失敗")
    print(f"\n  本次運單號：")
    for customer, waybill, pdf in completed_orders:
        print(f"    {customer}  →  {waybill}")
    print(f"\n  📊 Excel 追蹤表：{EXCEL_PATH}")
    print(f"{'='*60}")

    # ══ py6：自動重新列印所有運單（同一 Chrome session，無需確認）══════════════
    if _session:
        print("\n" + "="*60)
        print("  自動開始重新列印運單（py6）")
        print("="*60)
        reprint_ok = reprint_fail = 0
        for _entry in _session:
            _ok = reprint_one_waybill(ctx, _entry)
            if _ok: reprint_ok += 1
            else:   reprint_fail += 1
            time.sleep(1)
        print(f"\n  重印完成：✅ {reprint_ok} 成功  ❌ {reprint_fail} 失敗")

    # ══ 核對每個訂單 folder 是否有齊 4 個檔案，缺就補下載 ══════════════════════
    print("\n" + "="*60)
    print("  [核對] 檢查每個訂單 folder 是否有齊 4 個檔案")
    print("="*60)
    all_ok = True
    for _customer, _waybill, _pdf_path in completed_orders:
        _save_dir  = os.path.dirname(_pdf_path)
        _file_base = os.path.splitext(os.path.basename(_pdf_path))[0]
        # 從 file_base 提取 pos_order_no（以 ORD- 開頭的部分）
        _pos_order_no = next((p for p in _file_base.split("_") if p.startswith("ORD-")), None)

        _files = os.listdir(_save_dir) if os.path.exists(_save_dir) else []
        _has_ticket  = any(f.endswith((".pdf",".png")) and "_運單" not in f
                           and "_收貨明細" not in f and "_清關" not in f
                           and "_清單" not in f for f in _files)
        _has_waybill = any("_運單" in f and f.endswith(".pdf") for f in _files)
        _has_receipt = any("_收貨明細" in f for f in _files)
        _has_customs = any("_清關" in f for f in _files)

        _missing = []
        if not _has_ticket:  _missing.append("小票")
        if not _has_waybill: _missing.append("運單")
        if not _has_receipt: _missing.append("收貨明細")
        if not _has_customs: _missing.append("清關")

        if not _missing:
            print(f"  [OK] {_customer} ({_pos_order_no}) — 4/4 齊全")
        else:
            all_ok = False
            print(f"  [缺] {_customer} ({_pos_order_no}) — 缺少: {', '.join(_missing)}")
            # 如果缺收貨明細或清關，自動返 POS 補下載
            _need_pos = [m for m in _missing if m in ("收貨明細", "清關")]
            if _need_pos and _pos_order_no:
                print(f"    -> 自動補下載: {', '.join(_need_pos)}")
                try:
                    download_pos_word(ctx, _pos_order_no, _save_dir, _customer)
                except Exception as _re:
                    print(f"    -> 補下載失敗: {_re}")
            if "運單" in _missing:
                print(f"    -> 運單缺失，請手動跑 fix_waybill_v6.py 補印")
            if "小票" in _missing:
                print(f"    -> 小票缺失，請手動補印")

    if all_ok:
        print("\n  [完成] 所有訂單檔案齊全！")
    else:
        print("\n  [完成] 補下載已完成，請再核對上方缺失項目")
    print("="*60)

    input("\n按 Enter 關閉瀏覽器…")
    ctx.close()
