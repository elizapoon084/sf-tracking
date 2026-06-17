# -*- coding: utf-8 -*-
"""
generate_packing_list.py
========================
根據裝箱資料自動生成「万里交货清单」Excel 檔案（COWORK 新格式）。

使用方法：
  直接在 BOX_DATA 填入資料，然後按 Ctrl+F5 跑。
  或：python generate_packing_list.py

BOX_DATA 格式：
  每箱 = {"items": [{"code": "1084083", "qty": 42}, ...]}
  PV / 運費 自動從 data/cowork_prices.json 讀取，毋須手填。
  如需覆蓋：可在 item 層加 "pv": 610, "freight": 21

輸出：自動存到桌面「裝箱清單給珠海」資料夾。
"""

import os
import sys
import json
from datetime import date
from copy import copy

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# ── 路徑 ──────────────────────────────────────────────────────────────────────
TEMPLATE_PATH   = r"C:\Users\user\Desktop\順丰E順递\裝箱單給珠海模版\万里交货清单 31-5-2026 -肥仔運費-更正.xlsx"
COWORK_PRICES   = r"C:\Users\user\Desktop\順丰E順递\data\cowork_prices.json"
MANLEE_JSON     = r"C:\Users\user\Desktop\manlee_data.json"
OUTPUT_DIR      = r"C:\Users\user\Desktop\裝箱清單給珠海"


# ══════════════════════════════════════════════════════════════════════════════
# ★ 在這裡填入裝箱資料 ★
# ══════════════════════════════════════════════════════════════════════════════
BOX_DATA = [
    # 格式：{"items": [{"code":"代碼", "qty":數量}, ...]}
    # PV 和運費自動填入，毋須手填。
    # ── 每次跑腳本前更換以下資料 ─────────────────────────────────────────────
    # 第①箱 85件
    {"items": [{"code":"1000458","qty":40},{"code":"1000044","qty":20},{"code":"1000046","qty":10},{"code":"0100777","qty":1},{"code":"1084075","qty":4},{"code":"0700694","qty":4},{"code":"0300283","qty":1},{"code":"1084096","qty":5}]},
    # 第②箱 45件
    {"items": [{"code":"1084065","qty":29},{"code":"1084100","qty":16}]},
    # 第③箱 68件
    {"items": [{"code":"1084041","qty":67},{"code":"1084074","qty":1}]},
    # 第④箱 75件
    {"items": [{"code":"0300525","qty":42},{"code":"0300442","qty":5},{"code":"0300534","qty":8},{"code":"0300535","qty":8},{"code":"1084100","qty":2},{"code":"1084069","qty":8},{"code":"1084043","qty":2}]},
    # 第⑤箱 43件
    {"items": [{"code":"1084067","qty":39},{"code":"1084095","qty":4}]},
    # 第⑥箱 51件
    {"items": [{"code":"1084080","qty":26},{"code":"1084064","qty":20},{"code":"1084095","qty":5}]},
    # 第⑦箱 21件
    {"items": [{"code":"1084084","qty":12},{"code":"1084092","qty":9}]},
    # 第⑧箱 27件
    {"items": [{"code":"1084092","qty":5},{"code":"1084086","qty":15},{"code":"1084074","qty":1},{"code":"1084110","qty":5},{"code":"1084083","qty":1}]},
    # 第⑨箱 25件
    {"items": [{"code":"1084110","qty":12},{"code":"0300774","qty":1},{"code":"1084095","qty":4},{"code":"0100434","qty":4},{"code":"0100433","qty":4}]},
    # 第⑩箱 19件
    {"items": [{"code":"1084095","qty":17},{"code":"1084082","qty":2}]},
    # 第⑪箱 62件
    {"items": [{"code":"1084082","qty":45},{"code":"0300423","qty":7},{"code":"0300530","qty":10}]},
    # 第⑫箱 54件
    {"items": [{"code":"0300530","qty":22},{"code":"1084095","qty":3},{"code":"0300406","qty":16},{"code":"0300405","qty":12},{"code":"1000046","qty":1}]},
    # 第⑬箱 58件
    {"items": [{"code":"0300407","qty":29},{"code":"0300405","qty":9},{"code":"0300409","qty":20}]},
    # 第⑭箱 67件
    {"items": [{"code":"0300409","qty":7},{"code":"0300408","qty":18},{"code":"0300410","qty":16},{"code":"1084066","qty":26}]},
    # 第⑮箱 65件
    {"items": [{"code":"1084063","qty":40},{"code":"1084066","qty":25}]},
    # 第⑯箱 30件（精油留白）
    {"items": [{"code":"1084084","qty":28},{"code":"精油","qty":2,"name":"精油","pv":None,"freight":None}]},
    # 第⑰箱 42件
    {"items": [{"code":"1084083","qty":42}]},
    # 第⑱箱 32件
    {"items": [{"code":"1084093","qty":32}]},
    # 第⑲箱 9件
    {"items": [{"code":"1084085","qty":9}]},
    # 第⑳箱 72件
    {"items": [{"code":"1084064","qty":72}]},
    # 第㉑箱 74件
    {"items": [{"code":"1084080","qty":74}]},
    # 第㉒箱 32件
    {"items": [{"code":"1084092","qty":32}]},
    # 第㉓箱 36件
    {"items": [{"code":"1084043","qty":36}]},
    # 第㉔箱 36件
    {"items": [{"code":"1084043","qty":36}]},
    # 第㉕箱 80件
    {"items": [{"code":"1084082","qty":80}]},
    # 第㉖箱 36件
    {"items": [{"code":"1084043","qty":36}]},
    # 第㉗箱 70件
    {"items": [{"code":"1084067","qty":70}]},
    # 第㉘箱 70件
    {"items": [{"code":"1084065","qty":70}]},
    # 第㉙箱 42件
    {"items": [{"code":"1084083","qty":42}]},
    # 第㉚箱 12件
    {"items": [{"code":"1084085","qty":12}]},
    # 第㉛箱 36件
    {"items": [{"code":"1084043","qty":36}]},
    # 第㉜箱 17件
    {"items": [{"code":"0100777","qty":1},{"code":"1084093","qty":16}]},
    # 第㉝箱 74件
    {"items": [{"code":"1084080","qty":74}]},
    # 第㉞箱 12件
    {"items": [{"code":"1084085","qty":12}]},
    # 第㉟箱 12件
    {"items": [{"code":"1084085","qty":12}]},
    # 第㊱箱 36件
    {"items": [{"code":"1084043","qty":36}]},
    # 第㊲箱 36件
    {"items": [{"code":"1084043","qty":36}]},
]
# ══════════════════════════════════════════════════════════════════════════════


def load_db() -> dict:
    """載入 COWORK 定價 + manlee_data 產品名 → {code: {name, pv, freight}}"""
    db = {}

    # 1. cowork_prices.json（PV + 運費 + 名稱）
    try:
        raw = json.load(open(COWORK_PRICES, encoding="utf-8"))
        for code, info in raw.items():
            db[code] = {
                "name":    info.get("name", ""),
                "pv":      info.get("pv") or None,
                "freight": info.get("freight") or None,
            }
    except Exception as e:
        print(f"  ⚠️  cowork_prices.json 載入失敗: {e}")

    # 2. manlee_data.json（補充未收錄的產品名稱）
    try:
        raw2 = json.load(open(MANLEE_JSON, encoding="utf-8"))
        products = raw2.get("products", raw2)
        if isinstance(products, dict):
            products = list(products.values())
        for p in products:
            code = str(p.get("code", "")).strip()
            if code and code not in db:
                db[code] = {"name": p.get("name", ""), "pv": None, "freight": None}
            elif code and not db[code]["name"]:
                db[code]["name"] = p.get("name", "")
    except Exception as e:
        print(f"  ⚠️  manlee_data.json 載入失敗: {e}")

    print(f"  📦 產品資料庫：{len(db)} 個產品（{sum(1 for v in db.values() if v['pv'])} 個有COWORK定價）")
    return db


def generate(box_data: list, title: str = None, output_path: str = None):
    db = load_db()
    today_str = date.today().strftime("%Y-%m-%d")
    if not title:
        title = f"万里交货清单 {today_str}"

    # ── 複製模版並清除舊資料 ──────────────────────────────────────────────────
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.worksheets[0]

    # 清除所有合併儲存格（避免殘留 merge 令箱號寫唔入去）
    ws.merged_cells.ranges.clear() if hasattr(ws.merged_cells, 'ranges') else None
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

    last_row = ws.max_row
    for r in range(last_row, 2, -1):
        ws.delete_rows(r)

    ws.cell(1, 2).value = title

    # ── 寫入資料 ──────────────────────────────────────────────────────────────
    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row   = 3
    total_qty_sum = 0
    warnings      = []

    for box_idx, box in enumerate(box_data):
        box_no    = box_idx + 1
        items     = box["items"]
        total_qty = sum(it["qty"] for it in items)
        total_qty_sum += total_qty

        for item_idx, item in enumerate(items):
            code    = str(item.get("code", "")).strip()
            qty     = item.get("qty", 0)
            prod    = db.get(code, {})
            name    = prod.get("name", "")
            pv      = item.get("pv", prod.get("pv"))      # item 層覆蓋優先
            freight = item.get("freight", prod.get("freight"))

            r = current_row

            # A: 箱號（只第一行填）
            ws.cell(r, 1).value = box_no if item_idx == 0 else None
            # B: "1箱"
            ws.cell(r, 2).value = "1箱" if item_idx == 0 else None
            # C: "(XX個1箱)"
            ws.cell(r, 3).value = f"({total_qty}個1箱)" if item_idx == 0 else None
            # D: 產品代碼
            ws.cell(r, 4).value = code
            # E: 產品名稱
            ws.cell(r, 5).value = name
            # F: PV
            ws.cell(r, 6).value = pv
            # G: 數量
            ws.cell(r, 7).value = qty
            # H: 運費
            ws.cell(r, 8).value = freight
            # I: 總計 = 運費 × 數量
            ws.cell(r, 9).value = f"=H{r}*G{r}"

            for col in range(1, 10):
                cell = ws.cell(r, col)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if col in (1, 2, 3, 7, 8, 9) else "left",
                    vertical="center",
                )

            if not name and code:
                warnings.append(f"代碼 {code} 搵唔到名字")
            if pv is None and code:
                warnings.append(f"代碼 {code} 無COWORK定價（PV留空）")

            current_row += 1

    # ── 合計行 ────────────────────────────────────────────────────────────────
    r = current_row
    ws.cell(r, 7).value = f"=SUM(G3:G{r-1})"
    ws.cell(r, 9).value = f"=SUM(I3:I{r-1})"
    for col in [7, 9]:
        ws.cell(r, col).font   = Font(bold=True)
        ws.cell(r, col).border = border

    # ── 儲存 ──────────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if not output_path:
        output_path = os.path.join(OUTPUT_DIR, f"万里交货清单 {today_str}.xlsx")

    wb.save(output_path)

    for w in warnings:
        print(f"  ⚠️  {w}")
    print(f"\n  ✅ 已儲存：{output_path}")
    print(f"  📦 共 {len(box_data)} 箱  {total_qty_sum} 件")

    try:
        os.startfile(output_path)
    except Exception:
        pass

    return output_path


if __name__ == "__main__":
    print("=" * 55)
    print("  万里交货清单 自動生成器（COWORK 新格式）")
    print("=" * 55)
    generate(BOX_DATA)
    input("\n按 Enter 結束...")
