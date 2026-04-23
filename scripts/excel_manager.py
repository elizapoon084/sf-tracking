# -*- coding: utf-8 -*-
import os
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from config import (
    EXCEL_PATH, EXCEL_SHEET, EXCEL_HEADERS,
    COL_DATE, COL_NAME, COL_POS_ORDER, COL_WAYBILL,
    COL_RECIPIENT, COL_PHONE, COL_ADDRESS,
    COL_ITEMS, COL_QTY, COL_VIP_TOTAL,
    COL_PAYMENT, COL_FREIGHT, COL_STATUS,
    COL_STATUS_TIME, COL_ANOMALY, COL_PDF_PATH, COL_NOTES, COL_TAX,
    ANOMALY_KEYWORDS,
)
from logger import get_logger

log = get_logger(__name__)

_RED_FILL    = PatternFill("solid", fgColor="FF0000")
_HEADER_FONT = Font(bold=True)
_TOTAL_COLS  = len(EXCEL_HEADERS)


class ExcelManager:
    def __init__(self, path: str = EXCEL_PATH):
        self.path = path
        os.makedirs(os.path.dirname(path), exist_ok=True)
        self._load_or_create()

    # ── internal ──────────────────────────────────────────────────────────────

    def _load_or_create(self) -> None:
        if os.path.exists(self.path):
            self.wb = openpyxl.load_workbook(self.path)
            if EXCEL_SHEET not in self.wb.sheetnames:
                self.ws = self.wb.create_sheet(EXCEL_SHEET)
                self._write_header()
            else:
                self.ws = self.wb[EXCEL_SHEET]
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = EXCEL_SHEET
            self._write_header()
            self._auto_column_widths()
            self.save()
            log.info("Created new tracking.xlsx")

    def _write_header(self) -> None:
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT

    def _auto_column_widths(self) -> None:
        widths = [12, 10, 14, 16, 10, 14, 30, 28, 6, 12,
                  12, 10, 14, 18, 10, 40, 20, 10]  # last = 稅金
        for i, w in enumerate(widths, start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = w

    def save(self) -> None:
        self.wb.save(self.path)

    # ── public API ────────────────────────────────────────────────────────────

    def append_order(self, order: dict, pos_order_no: str = "",
                     payment: str = "到付", pdf_path: str = "") -> int:
        """Append a new row. Returns the row number written."""
        items_summary = ", ".join(
            f"{it.get('name', it['sku'])}×{it['qty']}" for it in order["items"]
        )
        total_qty = sum(it["qty"] for it in order["items"])

        row = self.ws.max_row + 1
        data = [""] * _TOTAL_COLS
        data[COL_DATE - 1]        = order["date"]
        data[COL_NAME - 1]        = order["name"]
        data[COL_POS_ORDER - 1]   = pos_order_no
        data[COL_WAYBILL - 1]     = ""
        data[COL_RECIPIENT - 1]   = order["name"]
        data[COL_PHONE - 1]       = order["phone"]
        data[COL_ADDRESS - 1]     = order["address"]
        data[COL_ITEMS - 1]       = items_summary
        data[COL_QTY - 1]         = total_qty
        data[COL_VIP_TOTAL - 1]   = order["total"]
        data[COL_PAYMENT - 1]     = payment
        data[COL_FREIGHT - 1]     = ""
        data[COL_STATUS - 1]      = "待寄出"
        data[COL_STATUS_TIME - 1] = ""
        data[COL_ANOMALY - 1]     = ""
        data[COL_PDF_PATH - 1]    = pdf_path
        data[COL_NOTES - 1]       = order.get("notes", "")

        for col, value in enumerate(data, start=1):
            self.ws.cell(row=row, column=col, value=value)

        self.save()
        log.info("Excel row %d appended for %s", row, order["name"])
        return row

    def update_waybill(self, row: int, waybill: str) -> None:
        self.ws.cell(row=row, column=COL_WAYBILL, value=waybill)
        if waybill:
            self.ws.cell(row=row, column=COL_STATUS, value="待派送")
        self.save()

    def update_pos_order(self, row: int, pos_order_no: str) -> None:
        self.ws.cell(row=row, column=COL_POS_ORDER, value=pos_order_no)
        self.save()

    def update_status(self, waybill: str, status: str,
                      freight: str = "", status_time: str = "") -> None:
        row = self.find_row_by_waybill(waybill)
        if row is None:
            log.warning("Waybill %s not found in Excel", waybill)
            return

        self.ws.cell(row=row, column=COL_STATUS, value=status)
        if freight:
            self.ws.cell(row=row, column=COL_FREIGHT, value=freight)
        if status_time:
            self.ws.cell(row=row, column=COL_STATUS_TIME, value=status_time)

        is_anomaly = any(kw in status for kw in ANOMALY_KEYWORDS)
        if is_anomaly:
            self._highlight_row_red(row)
            self.ws.cell(row=row, column=COL_ANOMALY, value="⚠️")
        self.save()

    def find_row_by_waybill(self, waybill: str) -> Optional[int]:
        for row in self.ws.iter_rows(min_row=2):
            cell = row[COL_WAYBILL - 1]
            if cell.value and str(cell.value).strip() == waybill.strip():
                return cell.row
        return None

    def get_all_waybills(self) -> list:
        """Returns [(row_number, waybill_str)] for rows with a waybill."""
        result = []
        for row in self.ws.iter_rows(min_row=2):
            waybill = row[COL_WAYBILL - 1].value
            status  = row[COL_STATUS - 1].value or ""
            if waybill and "簽收" not in status:
                result.append((row[0].row, str(waybill).strip()))
        return result

    def get_recent_rows(self, n: int = 10) -> list:
        """Returns last n data rows as list of dicts for GUI display."""
        rows = []
        for r in self.ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in r):
                rows.append(r)
        return rows[-n:]

    def append_from_sf(self, waybill: str, sf_info: dict) -> int:
        """Add a new row sourced from SF scrape (no POS data available)."""
        from datetime import date
        row = self.ws.max_row + 1
        data = [""] * _TOTAL_COLS
        # Use date from SF list page (status date) as 寄出時間 if available
        sf_date = sf_info.get("date", "")
        data[COL_DATE - 1]        = sf_date[:10] if sf_date else str(date.today())
        data[COL_WAYBILL - 1]     = waybill
        data[COL_STATUS - 1]      = sf_info.get("status", "")
        data[COL_STATUS_TIME - 1] = sf_date
        for col, value in enumerate(data, start=1):
            self.ws.cell(row=row, column=col, value=value)
        self.save()
        log.info("Appended SF waybill row %d: %s", row, waybill)
        return row

    def update_receipt_detail(self, waybill: str, detail: dict) -> None:
        """Write 電子存根 fields into the matching Excel row."""
        row = self.find_row_by_waybill(waybill)
        if row is None:
            log.warning("Waybill %s not found for receipt update", waybill)
            return

        def _set(col, key):
            v = detail.get(key)
            if v:
                self.ws.cell(row=row, column=col, value=v)

        _set(COL_RECIPIENT,    "recipient_name")
        # Fill COL_NAME if empty or still has wrong sender name "Eliza poon"
        if detail.get("recipient_name"):
            existing_name = str(self.ws.cell(row=row, column=COL_NAME).value or "").strip()
            if existing_name in ("", "None", "nan", "Eliza poon"):
                self.ws.cell(row=row, column=COL_NAME, value=detail["recipient_name"])
        _set(COL_PHONE,        "recipient_phone")
        _set(COL_ADDRESS,      "recipient_address")
        _set(COL_ITEMS,        "items")
        _set(COL_FREIGHT,      "freight")
        _set(COL_STATUS_TIME,  "delivery_time")

        if detail.get("pieces"):
            try:
                self.ws.cell(row=row, column=COL_QTY, value=int(detail["pieces"]))
            except (ValueError, TypeError):
                pass

        # Pack remaining fields into notes
        extra_parts = []
        for label, key in [
            ("實重",   "actual_weight"),
            ("計重",   "chargeable_weight"),
            ("類型",   "product_type"),
            ("付款",   "payment"),
            ("收件員", "delivery_person"),
        ]:
            if detail.get(key):
                extra_parts.append(f"{label}:{detail[key]}")

        if extra_parts:
            existing = str(self.ws.cell(row=row, column=COL_NOTES).value or "")
            addition = " | ".join(extra_parts)
            self.ws.cell(row=row, column=COL_NOTES,
                         value=(existing + " | " + addition).lstrip(" | "))

        self.save()
        log.info("Receipt detail saved for %s", waybill)

    def _highlight_row_red(self, row: int) -> None:
        for col in range(1, _TOTAL_COLS + 1):
            self.ws.cell(row=row, column=col).fill = _RED_FILL
