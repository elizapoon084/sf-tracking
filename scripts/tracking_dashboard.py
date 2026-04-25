# -*- coding: utf-8 -*-
# tracking_dashboard.py — 順豐寄件追蹤儀表板
# 執行: python -m streamlit run scripts/tracking_dashboard.py --server.port 8502
import os
import sys
import time
import subprocess

import base64
import io
import json
from collections import defaultdict
from datetime import date as _date
import streamlit as st
import pandas as pd
import openpyxl

# ── 雲端 / 本地 自動偵測 ───────────────────────────────────────────────────────
# 喺 Streamlit Cloud 上，用相對路徑讀 Excel；本地用 config.py 的絕對路徑
_IS_CLOUD = os.environ.get("STREAMLIT_SHARING_MODE") or not os.path.exists(
    r"C:\Users\user\Desktop\順丰E順递"
)

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from config import (
        EXCEL_PATH as _LOCAL_EXCEL_PATH, EXCEL_SHEET,
        COL_DATE, COL_NAME, COL_WAYBILL,
        COL_RECIPIENT, COL_PHONE, COL_ADDRESS,
        COL_ITEMS, COL_QTY,
        COL_STATUS, COL_STATUS_TIME,
        COL_FREIGHT, COL_NOTES, COL_TAX,
        COL_PDF_PATH,
        ANOMALY_KEYWORDS,
    )
except Exception:
    # 雲端環境沒有 config.py，使用預設值
    _LOCAL_EXCEL_PATH = ""
    EXCEL_SHEET   = "追蹤表"
    COL_DATE=1; COL_NAME=2; COL_WAYBILL=4; COL_RECIPIENT=5; COL_PHONE=6
    COL_ADDRESS=7; COL_ITEMS=8; COL_QTY=9; COL_STATUS=13; COL_STATUS_TIME=14
    COL_FREIGHT=12; COL_NOTES=17; COL_TAX=18; COL_PDF_PATH=16
    ANOMALY_KEYWORDS = ["退回","異常","卡關","攔截","問題件"]

# 雲端用相對路徑，本地用 config.py 的絕對路徑
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_CLOUD_EXCEL = os.path.join(_SCRIPT_DIR, "..", "data", "tracking.xlsx")
EXCEL_PATH = _CLOUD_EXCEL if _IS_CLOUD else _LOCAL_EXCEL_PATH

SF_PUBLIC_TRACK = "https://www.sf-express.com/cn/sc/dynamic_function/waybill/#search/bill-number/{}"


# ── 稅金寫回 Excel ─────────────────────────────────────────────────────────────

def _push_to_github() -> bool:
    """Commit updated tracking.xlsx to GitHub (used in cloud mode)."""
    try:
        import base64, requests as _req
        token   = st.secrets.get("GITHUB_TOKEN", "")
        repo    = st.secrets.get("GITHUB_REPO", "elizapoon084/sf-tracking")
        gh_path = st.secrets.get("GITHUB_FILE_PATH", "data/tracking.xlsx")
        if not token:
            return False
        headers = {"Authorization": f"token {token}",
                   "Accept": "application/vnd.github.v3+json"}
        url = f"https://api.github.com/repos/{repo}/contents/{gh_path}"
        sha = _req.get(url, headers=headers, timeout=15).json().get("sha", "")
        with open(EXCEL_PATH, "rb") as f:
            content = base64.b64encode(f.read()).decode()
        from datetime import datetime
        r = _req.put(url, headers=headers, timeout=30, json={
            "message": f"cloud: 更新稅金 {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "content": content, "sha": sha,
        })
        return r.status_code in (200, 201)
    except Exception as e:
        st.warning(f"⚠️ GitHub 同步失敗：{e}")
        return False


def _save_tax_values(changed_df: pd.DataFrame) -> int:
    """Write updated tax values back to tracking.xlsx. Returns count saved."""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[EXCEL_SHEET]
        saved = 0
        for _, row in changed_df.iterrows():
            waybill = str(row.get("運單號", "")).strip()
            tax     = row.get("稅金(HKD)", None)
            if not waybill or waybill in ("None", "nan", ""):
                continue
            for excel_row in ws.iter_rows(min_row=2):
                wb_cell = excel_row[COL_WAYBILL - 1]
                if str(wb_cell.value or "").strip() == waybill:
                    excel_row[COL_TAX - 1].value = float(tax) if tax not in (None, "", "nan") else None
                    saved += 1
                    break
        wb.save(EXCEL_PATH)
        if _IS_CLOUD:
            ok = _push_to_github()
            if ok:
                st.success("✅ 稅金已同步到 GitHub，本地 Excel 將在下次排程時更新")
            else:
                st.warning("⚠️ 本地儲存成功，但 GitHub 同步失敗，請檢查 Secrets 設定")
        return saved
    except Exception as e:
        st.error(f"儲存稅金失敗：{e}")
        return 0


# ── Status badge colours ───────────────────────────────────────────────────────
_STATUS_COLOR = {
    "已簽收":   ("#27ae60", "✅"),
    "派送中":   ("#2980b9", "🚚"),
    "待派送":   ("#8e44ad", "📬"),
    "待簽收":   ("#8e44ad", "📬"),
    "運送中":   ("#2980b9", "🚚"),
    "攬收成功": ("#16a085", "📦"),
    "攬收":     ("#16a085", "📦"),
    "已發出":   ("#16a085", "📤"),
    "待寄出":   ("#95a5a6", "⏳"),
    "已取消":   ("#bdc3c7", "❌"),
    "退回":     ("#e74c3c", "↩️"),
    "異常":     ("#e74c3c", "⚠️"),
    "卡關":     ("#e74c3c", "🚫"),
    "攔截":     ("#e74c3c", "🚫"),
    "問題件":   ("#e74c3c", "⚠️"),
    "更新失敗": ("#e67e22", "❓"),
    "查詢不到": ("#e67e22", "❓"),
    "狀態不明": ("#e67e22", "❓"),
}

def _badge(status: str) -> str:
    s = str(status)
    color, icon = "#95a5a6", ""
    for kw, (c, i) in _STATUS_COLOR.items():
        if kw in s:
            color, icon = c, i
            break
    return (
        f'<span style="background:{color};color:#fff;padding:3px 10px;'
        f'border-radius:12px;font-size:12px;font-weight:600;white-space:nowrap;">'
        f'{icon} {s}</span>'
    )

def _val(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:   # NaN check (NaN != NaN)
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none") else s

def _esc(s) -> str:
    v = _val(s)
    return v.replace('"', "&quot;").replace("<", "&lt;").replace(">", "&gt;")

_HK_KW  = ["香港", "九龍", "新界", "kowloon", "hong kong", "東區", "中區",
            "灣仔", "油麻地", "旺角", "屯門", "沙田", "大埔", "將軍澳",
            "柴灣", "觀塘", "荃灣", "元朗", "上水", "粉嶺", "葵涌"]
_CN_KW  = ["廣東省", "深圳市", "廣州市", "北京市", "上海市", "福建省", "浙江省", "江蘇省",
            "湖南省", "四川省", "重慶市", "天津市", "成都市", "杭州市", "南京市", "武漢市",
            "東莞市", "中山市", "佛山市", "珠海市", "惠州市", "廈門市"]

def _dest(addr: str) -> str:
    a = addr.lower()
    for kw in _HK_KW:
        if kw.lower() in a:
            return "🇭🇰 香港"
    for kw in _CN_KW:
        if kw.lower() in a:
            return "🇨🇳 中國"
    return ""


# ── Data loading ──────────────────────────────────────────────────────────────

_SCRIPT_DIR2 = os.path.dirname(os.path.abspath(__file__))
_PRODUCTS_JSON = (
    os.path.join(_SCRIPT_DIR2, "..", "data", "products.json")
    if _IS_CLOUD else
    r"C:\Users\user\Desktop\順丰E順递\data\products.json"
)

@st.cache_data(ttl=300)
def load_name_to_sku() -> dict:
    """產品名稱 → SKU 反查表"""
    try:
        with open(_PRODUCTS_JSON, encoding="utf-8") as f:
            data = json.load(f)
        return {info.get("name", ""): sku for sku, info in data.items() if info.get("name")}
    except Exception:
        return {}


try:
    import zhconv as _zhconv
    def _to_hant(s: str) -> str:
        """Convert Simplified → Traditional for name comparison/grouping."""
        return _zhconv.convert(s, "zh-hant") if s else s
except ImportError:
    def _to_hant(s: str) -> str:
        return s


def _parse_item(item_str: str) -> tuple:
    """Parse '[SKU] Name×qty' or 'Name×qty' → (sku, name, qty).
    Supports both × and x as quantity separator."""
    s = item_str.strip()
    sku = ""
    if s.startswith("[") and "]" in s:
        end = s.index("]")
        sku = s[1:end].strip()
        s = s[end + 1:].strip()
    sep = "×" if "×" in s else ("x" if "x" in s else "")
    if sep:
        nm, q = s.rsplit(sep, 1)
        try:
            return sku, nm.strip(), int(q.strip())
        except ValueError:
            return sku, nm.strip(), 1
    return sku, s, 0

@st.cache_data(ttl=30)
def load_orders() -> pd.DataFrame:
    if not os.path.exists(EXCEL_PATH):
        return pd.DataFrame()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb[EXCEL_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return pd.DataFrame()
        headers = list(rows[0])
        data = [list(r) for r in rows[1:] if any(v is not None for v in r)]
        return pd.DataFrame(data, columns=headers)
    except Exception as e:
        st.error(f"讀取 Excel 失敗：{e}")
        return pd.DataFrame()


# ── Status refresh ────────────────────────────────────────────────────────────

def _run_status_update() -> str:
    cli = os.path.join(os.path.dirname(os.path.abspath(__file__)), "update_status_cli.py")
    result = subprocess.run(
        [sys.executable, cli],
        capture_output=True, text=True, encoding="utf-8", timeout=300,
    )
    return (result.stdout or "") + (result.stderr or "")


# ── Main layout ───────────────────────────────────────────────────────────────

def main():
    st.set_page_config(page_title="順豐寄件追蹤", page_icon="📦",
                       layout="wide", initial_sidebar_state="expanded")

    st.markdown("""
    <style>
    .block-container { padding-top: 1.2rem; padding-bottom: 1rem; }
    div[data-testid="metric-container"] {
        background: #f8f9fa; border-radius: 10px;
        padding: 12px 16px; border: 1px solid #e9ecef;
    }
    thead tr th { position: sticky; top: 0; z-index: 1; }
    </style>
    """, unsafe_allow_html=True)

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## 📦 順豐追蹤")
        if _IS_CLOUD:
            st.info("☁️ 雲端唯讀模式\n\n資料由本地電腦定時同步，查件請點擊運單號。")
        st.divider()

        if not _IS_CLOUD:
            if st.button("🔄 向順豐查詢最新狀態", use_container_width=True, type="primary"):
                with st.spinner("正在查詢順豐，請稍候（1-2 分鐘）…"):
                    out = _run_status_update()
                st.success("✅ 狀態更新完成")
                if out:
                    with st.expander("查看輸出記錄"):
                        st.code(out)
                st.cache_data.clear()
                st.rerun()

        st.divider()
        st.markdown("#### 🔍 篩選")
        status_options = [
            "全部", "待寄出", "攬收成功", "運送中",
            "待派送", "待簽收", "派送中", "已簽收", "退回", "異常", "問題件",
        ]
        sel_status = st.selectbox("狀態篩選", status_options, label_visibility="collapsed")
        dest_options = ["全部目的地", "🇭🇰 香港", "🇨🇳 中國"]
        sel_dest   = st.selectbox("目的地", dest_options, label_visibility="collapsed")
        search_kw  = st.text_input("🔎 搜尋客人名 / 運單號 / 地址")

        st.divider()
        st.caption("💡 點擊運單號可在順豐網站查件")
        st.caption("💡 「全部（顯示活躍）」隱藏已取消訂單")

    # ── Load data ─────────────────────────────────────────────────────────────
    df = load_orders()
    name_to_sku = load_name_to_sku()   # 產品名 → SKU
    if df.empty:
        st.info("📭 暫無訂單記錄。")
        st.caption(f"Excel 路徑：{EXCEL_PATH}")
        return

    # 最新日期排最頂（按日期降序，日期相同保留原始倒序）
    date_col_raw = df.columns[COL_DATE - 1]
    df["_sort_date"] = pd.to_datetime(df[date_col_raw].astype(str).str[:8], format="%Y%m%d", errors="coerce")
    df = df.iloc[::-1].reset_index(drop=True)  # 先倒序（同日期內最新在頂）
    df = df.sort_values("_sort_date", ascending=False, kind="stable", na_position="last")
    df = df.drop(columns=["_sort_date"]).reset_index(drop=True)

    status_col    = df.columns[COL_STATUS     - 1]
    waybill_col   = df.columns[COL_WAYBILL    - 1]
    date_col      = df.columns[COL_DATE       - 1]
    name_col      = df.columns[COL_NAME       - 1]
    items_col     = df.columns[COL_ITEMS      - 1]
    qty_col       = df.columns[COL_QTY        - 1]
    addr_col      = df.columns[COL_ADDRESS    - 1]
    recipient_col = df.columns[COL_RECIPIENT  - 1]
    stime_col     = df.columns[COL_STATUS_TIME- 1]
    freight_col   = df.columns[COL_FREIGHT    - 1]
    notes_col     = df.columns[COL_NOTES      - 1]
    phone_col     = df.columns[COL_PHONE      - 1]
    # COL_TAX may not exist yet in older Excel files — guard with index check
    tax_col       = df.columns[COL_TAX - 1] if len(df.columns) >= COL_TAX else None

    s = df[status_col].astype(str)
    total     = len(df)
    delivered = int(s.str.contains("已簽收").sum())
    in_flight = int(s.str.contains("派送中|運送中|攬收|待派送|待簽收").sum())
    pending   = int(s.str.contains("待寄出").sum())
    cancelled = int(s.str.contains("已取消").sum())
    anomaly   = int(s.str.contains("|".join(ANOMALY_KEYWORDS)).sum())

    # ── Summary stats ─────────────────────────────────────────────────────────
    st.markdown("### 📊 概況")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("📦 總訂單",  total)
    c2.metric("✅ 已簽收",  delivered)
    c3.metric("🚚 運送中",  in_flight)
    c4.metric("⏳ 待寄出",  pending)
    c5.metric("❌ 已取消",  cancelled)
    c6.metric("⚠️ 異常",    anomaly,
              delta="需跟進" if anomaly else "正常",
              delta_color="inverse")

    st.divider()

    # ── Apply filters ─────────────────────────────────────────────────────────
    # Always remove 已取消 — cancelled orders never shown in table
    display = df[~df[status_col].astype(str).str.contains("已取消", na=False)].copy()

    if sel_status not in ("全部（顯示活躍）", "全部"):
        display = display[display[status_col].astype(str).str.contains(sel_status, na=False)]

    if sel_dest != "全部目的地":
        dest_key = "香港" if "香港" in sel_dest else "中國"
        display = display[display[addr_col].apply(
            lambda v: dest_key in _dest(_val(v)))]

    if search_kw:
        kw = search_kw.lower()
        display = display[display.apply(
            lambda r: kw in " ".join(_val(v) for v in r).lower(), axis=1)]

    st.caption(f"顯示 **{len(display)}** / {total} 條記錄")

    # ── Order table ───────────────────────────────────────────────────────────
    rows_html = []
    logistics_keys = []   # 用於製作物流單的選單
    for i, (_, row) in enumerate(display.iterrows()):
        wb_val      = _val(row[waybill_col])
        status      = _val(row[status_col]) or "—"
        name_v      = _to_hant(_val(row[name_col]))
        date_v      = _val(row[date_col])
        items_v     = _val(row[items_col])
        qty_v       = _val(row[qty_col])
        addr_v      = _val(row[addr_col])
        recipient_v = _val(row[recipient_col])
        stime_v     = _val(row[stime_col])
        freight_v   = _val(row[freight_col])
        notes_v     = _val(row[notes_col])
        tax_v       = _val(row[tax_col]) if tax_col is not None else ""

        is_anom     = any(kw in status for kw in ANOMALY_KEYWORDS)
        is_cancel   = "已取消" in status
        is_signed   = "已簽收" in status

        if is_anom:
            row_bg = "background:#fff5f5;"
        elif is_cancel:
            row_bg = "background:#fafafa;opacity:0.75;"
        elif i % 2 == 0:
            row_bg = "background:#ffffff;"
        else:
            row_bg = "background:#f8fafc;"

        # ── 運單號 cell ───────────────────────────────────────────────────────
        wb_cell = (
            f'<a href="{SF_PUBLIC_TRACK.format(wb_val)}" target="_blank" '
            f'style="color:#2980b9;font-family:monospace;font-size:12px;'
            f'text-decoration:none;font-weight:600;">{wb_val}</a>'
            if wb_val else '<span style="color:#ccc">—</span>'
        )

        # ── 貨品 cell（點擊展開詳情）─────────────────────────────────────────
        if items_v:
            preview = items_v[:45] + ("…" if len(items_v) > 45 else "")
            detail_lines = "".join(
                f'<div style="padding:2px 0;border-bottom:1px solid #eee;">• {_esc(it.strip())}</div>'
                for it in items_v.split(" / ") if it.strip()
            )
            items_cell = (
                f'<details style="cursor:pointer;min-width:180px;">'
                f'<summary style="color:#2980b9;font-size:12px;list-style:none;'
                f'cursor:pointer;outline:none;">▶ {_esc(preview)}</summary>'
                f'<div style="margin-top:6px;padding:8px 10px;background:#f0f8ff;'
                f'border-radius:6px;border-left:3px solid #2980b9;'
                f'font-size:12px;line-height:1.9;min-width:220px;">'
                f'{detail_lines}</div>'
                f'</details>'
            )
        else:
            items_cell = '<span style="color:#bbb;font-size:11px;">—</span>'

        # ── 地址 cell ─────────────────────────────────────────────────────────
        dest_tag = _dest(addr_v)
        dest_html = (
            f'<span style="font-size:10px;background:#ecf0f1;border-radius:4px;'
            f'padding:1px 5px;margin-right:4px;">{dest_tag}</span>'
            if dest_tag else ""
        )
        addr_cell = (
            f'{dest_html}<span title="{_esc(addr_v)}" style="cursor:help;font-size:12px;color:#555;">'
            f'{_esc(addr_v[:28])}{"…" if len(addr_v) > 28 else ""}</span>'
            if addr_v else '<span style="color:#bbb;font-size:11px;">—</span>'
        )

        # ── 客人名 cell (tooltip shows notes) ────────────────────────────────
        name_html = (
            f'<span title="{_esc(notes_v)}" style="cursor:help;font-weight:700;">{_esc(name_v)}</span>'
            if notes_v and name_v else
            f'<b>{_esc(name_v)}</b>' if name_v else
            '<span style="color:#bbb;font-size:11px;">—</span>'
        )

        # ── 狀態 cell — badge + 簽收時間 for signed items ────────────────────
        if is_signed and stime_v:
            status_cell = (
                f'{_badge(status)}<br>'
                f'<span style="font-size:10px;color:#27ae60;white-space:nowrap;">'
                f'🕐 {_esc(stime_v[:16])}</span>'
            )
        else:
            status_cell = _badge(status)

        # ── 電子存根 cell ─────────────────────────────────────────────────────
        if is_signed:
            if freight_v or recipient_v:
                receipt_lines = []
                if recipient_v:
                    receipt_lines.append(
                        f'<b style="color:#2c3e50;">收：</b>{_esc(recipient_v)}'
                    )
                # 費用合計
                if freight_v:
                    receipt_lines.append(
                        f'<b>費用合計：</b>'
                        f'<span style="color:#27ae60;font-weight:700;">HKD {_esc(freight_v)}</span>'
                    )
                # 產品類型、件數 from notes
                if notes_v:
                    for part in notes_v.split("|"):
                        part = part.strip()
                        if part.startswith("類型"):
                            receipt_lines.append(
                                f'<b>產品類型：</b>'
                                f'<span style="color:#8e44ad;">{_esc(part.replace("類型:","").strip())}</span>'
                            )
                # 件數
                if qty_v:
                    receipt_lines.append(f'<b>件數：</b>{_esc(qty_v)}')
                # 收件時間
                if stime_v:
                    receipt_lines.append(
                        f'<b>收件時間：</b>'
                        f'<span style="color:#2980b9;font-size:11px;">{_esc(stime_v[:19])}</span>'
                    )
                receipt_cell = "<br>".join(receipt_lines)
            else:
                receipt_cell = (
                    '<span style="color:#e67e22;font-size:11px;">⏳ 待取存根<br>'
                    '（下次更新時自動抓取）</span>'
                )
        else:
            receipt_cell = '<span style="color:#ddd;font-size:11px;">—</span>'

        tax_html = (
            f'<span style="color:#e67e22;font-weight:600;">HKD {_esc(tax_v)}</span>'
            if tax_v else
            '<span style="color:#ddd;font-size:11px;">—</span>'
        )

        lkey = f"#{i+1}  {_val(row[date_col])}  {_val(row[name_col])}  {_val(row[waybill_col])}"
        logistics_keys.append((lkey, row))

        rows_html.append(f"""
        <tr style="{row_bg}border-bottom:1px solid #e9ecef;">
          <td style="padding:8px 6px;text-align:center;color:#aaa;font-size:11px;white-space:nowrap;">#{i+1}</td>
          <td style="padding:8px 10px;white-space:nowrap;font-size:12px;color:#666;">{_esc(date_v)}</td>
          <td style="padding:8px 10px;font-size:14px;min-width:110px;">{name_html}</td>
          <td style="padding:8px 10px;">{wb_cell}</td>
          <td style="padding:8px 10px;">{items_cell}</td>
          <td style="padding:8px 10px;text-align:center;font-size:13px;">{_esc(qty_v)}</td>
          <td style="padding:8px 10px;">{addr_cell}</td>
          <td style="padding:8px 10px;text-align:center;">{status_cell}</td>
          <td style="padding:8px 10px;font-size:12px;line-height:1.6;">{receipt_cell}</td>
          <td style="padding:8px 10px;text-align:right;">{tax_html}</td>
        </tr>""")

    table_html = f"""
    <div style="overflow-x:auto;border-radius:10px;border:1px solid #e9ecef;
                box-shadow:0 1px 4px rgba(0,0,0,.06);">
    <table style="width:100%;border-collapse:collapse;font-size:13px;font-family:sans-serif;">
    <thead>
    <tr style="background:#34495e;color:#fff;font-size:12px;letter-spacing:.5px;">
      <th style="padding:10px 6px;text-align:center;font-weight:600;color:#aaa;">#</th>
      <th style="padding:10px 10px;text-align:left;font-weight:600;">寄出時間</th>
      <th style="padding:10px 10px;text-align:left;font-weight:600;min-width:110px;">客人</th>
      <th style="padding:10px 10px;text-align:left;font-weight:600;">運單號</th>
      <th style="padding:10px 10px;text-align:left;font-weight:600;">貨品</th>
      <th style="padding:10px 10px;text-align:center;font-weight:600;">件</th>
      <th style="padding:10px 10px;text-align:left;font-weight:600;">收件地址</th>
      <th style="padding:10px 10px;text-align:center;font-weight:600;">狀態</th>
      <th style="padding:10px 10px;text-align:left;font-weight:600;">電子存根</th>
      <th style="padding:10px 10px;text-align:right;font-weight:600;">稅金</th>
    </tr>
    </thead>
    <tbody>{''.join(rows_html)}</tbody>
    </table></div>"""

    st.markdown(table_html, unsafe_allow_html=True)

    # ── 訂單貨品詳情 ──────────────────────────────────────────────────────────
    st.divider()
    with st.expander("📋 點開查看訂單貨品詳情", expanded=False):
        order_keys = [
            f"{_val(r[date_col])}  {_val(r[name_col])}  {_val(r[waybill_col])}"
            for _, r in df.iterrows() if _val(r[waybill_col])
        ]
        if order_keys:
            pick = st.selectbox("選擇訂單", ["— 請選擇 —"] + order_keys, key="detail_pick")
            if pick != "— 請選擇 —":
                matched = df[df.apply(
                    lambda r: f"{_val(r[date_col])}  {_val(r[name_col])}  {_val(r[waybill_col])}" == pick,
                    axis=1)]
                if not matched.empty:
                    r = matched.iloc[0]
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(f"**客人：** {_val(r[name_col])}")
                        st.markdown(f"**運單號：** `{_val(r[waybill_col])}`")
                        st.markdown(f"**日期：** {_val(r[date_col])}")
                        st.markdown(f"**狀態：** {_val(r[status_col])}")
                    with c2:
                        st.markdown(f"**電話：** {_val(r[phone_col])}")
                        st.markdown(f"**地址：** {_val(r[addr_col])}")
                    st.markdown("**貨品清單：**")
                    for item in _val(r[items_col]).split(" / "):
                        item = item.strip()
                        if item:
                            sku_i, nm_i, qty_i = _parse_item(item)
                            sku_tag = f"<span style='color:#888;font-size:11px;'>[{sku_i}]</span> " if sku_i else ""
                            if qty_i:
                                st.markdown(
                                    f"　• {sku_tag}<b>{_esc(nm_i)}</b> &nbsp;×&nbsp; "
                                    f"<span style='color:#e74c3c;font-weight:700;'>{qty_i} 件</span>",
                                    unsafe_allow_html=True)
                            else:
                                st.markdown(f"　• {sku_tag}{_esc(nm_i)}", unsafe_allow_html=True)

    # ── 批次發貨清單 ──────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 📦 批次發貨清單")
    st.caption("選擇已簽收訂單，一鍵生成合併貨品清單並下載")

    signed_mask = df[status_col].astype(str).str.contains("已簽收", na=False)
    signed_df   = df[signed_mask].copy()

    if signed_df.empty:
        st.info("暫無已簽收訂單")
    else:
        # Normalise names: Simplified → Traditional for grouping
        signed_df["_name_hant"] = signed_df[name_col].apply(
            lambda v: _to_hant(_val(v)))
        customers = ["全部客人"] + sorted(
            [x for x in signed_df["_name_hant"].dropna().unique() if x])
        sel_cust = st.selectbox("篩選客人", customers, key="batch_cust")

        batch_pool = signed_df if sel_cust == "全部客人" else \
            signed_df[signed_df["_name_hant"] == sel_cust]

        order_labels = [
            f"{_val(r[date_col])}  {_to_hant(_val(r[name_col]))}  {_val(r[waybill_col])}"
            for _, r in batch_pool.iterrows()
        ]

        if order_labels:
            sel_orders = st.multiselect(
                f"勾選訂單（共 {len(order_labels)} 張已簽收）",
                order_labels, default=order_labels, key="batch_sel")

            if sel_orders and st.button("📊 生成合併發貨清單", type="primary"):
                totals = defaultdict(int)
                sku_map: dict[str, str] = {}
                included_rows = []
                for label in sel_orders:
                    for _, row in batch_pool.iterrows():
                        if f"{_val(row[date_col])}  {_to_hant(_val(row[name_col]))}  {_val(row[waybill_col])}" == label:
                            included_rows.append(row)
                            for item in _val(row[items_col]).split(" / "):
                                item = item.strip()
                                if not item:
                                    continue
                                sku_i, nm_i, qty_i = _parse_item(item)
                                if qty_i:
                                    totals[nm_i] += qty_i
                                    if nm_i not in sku_map:
                                        sku_map[nm_i] = sku_i or name_to_sku.get(nm_i, "")
                            break

                if totals:
                    total_kinds = len(totals)
                    total_qty   = sum(totals.values())
                    st.markdown(f"#### ✅ 合併結果：{total_kinds} 種貨品，共 {total_qty} 件")

                    result_rows = sorted(totals.items(), key=lambda x: x[0])
                    st.dataframe(
                        pd.DataFrame(
                            [{"編號": sku_map.get(nm, name_to_sku.get(nm, "")),
                              "貨品名稱": nm, "總數量": qty}
                             for nm, qty in result_rows],
                            columns=["編號", "貨品名稱", "總數量"]),
                        use_container_width=True, hide_index=True)

                    # ── 生成 Excel 下載 ──────────────────────────────────────
                    buf = io.BytesIO()
                    wb_out = openpyxl.Workbook()
                    ws_out = wb_out.active
                    ws_out.title = "發貨清單"
                    ws_out.append(["編號(SKU)", "貨品名稱", "總數量"])
                    for nm, qty in result_rows:
                        ws_out.append([sku_map.get(nm, name_to_sku.get(nm, "")), nm, qty])
                    ws_out.append([])
                    ws_out.append([f"共 {total_kinds} 種", f"合計 {total_qty} 件"])
                    ws_out.append([])
                    ws_out.append(["訂單明細"])
                    ws_out.append(["日期", "客人", "運單號", "貨品"])
                    for row in included_rows:
                        ws_out.append([
                            _val(row[date_col]), _val(row[name_col]),
                            _val(row[waybill_col]), _val(row[items_col]),
                        ])
                    wb_out.save(buf)
                    buf.seek(0)

                    fname = f"發貨清單_{sel_cust}_{_date.today().strftime('%Y%m%d')}.xlsx"
                    st.download_button(
                        "📥 下載 Excel 發貨清單",
                        data=buf, file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── 製作物流單 ────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 🚚 製作物流單")
    st.caption("按表格左側 # 號找到訂單，選擇貨品後匯出")

    if logistics_keys:
        lkey_labels = [k for k, _ in logistics_keys]
        sel_lkey = st.selectbox("選擇訂單", ["— 請選擇 —"] + lkey_labels, key="logistics_sel")

        if sel_lkey != "— 請選擇 —":
            sel_row = next(r for k, r in logistics_keys if k == sel_lkey)
            recip_v = _val(sel_row[recipient_col])
            ph_v    = _val(sel_row[phone_col])
            addr_v2 = _val(sel_row[addr_col])
            items_s = _val(sel_row[items_col])

            st.markdown(
                f"**收件人：** {recip_v} &nbsp;｜&nbsp; **電話：** {ph_v} &nbsp;｜&nbsp; **地址：** {addr_v2}")

            st.markdown("**選擇貨品及數量：**")
            parsed = []
            for it in items_s.split(" / "):
                it = it.strip()
                if not it:
                    continue
                if "×" in it:
                    nm, q = it.rsplit("×", 1)
                    try:
                        orig = int(q.strip())
                    except ValueError:
                        orig = 1
                    parsed.append((nm.strip(), orig))
                else:
                    parsed.append((it, 1))

            selected_items = []
            for idx2, (nm, orig) in enumerate(parsed):
                c1, c2, c3 = st.columns([0.04, 0.62, 0.34])
                with c1:
                    chk = st.checkbox("", value=True, key=f"lchk_{sel_lkey}_{idx2}")
                with c2:
                    st.markdown(f"**{nm}**")
                with c3:
                    qty2 = st.number_input(
                        "數量", min_value=0, max_value=orig * 5,
                        value=orig, step=1,
                        key=f"lqty_{sel_lkey}_{idx2}",
                        label_visibility="collapsed")
                if chk and qty2 > 0:
                    selected_items.append({"name": nm, "qty": qty2})

            if selected_items:
                if st.button("📥 匯出物流單 Excel", type="primary", key="logistics_export"):
                    buf2 = io.BytesIO()
                    wb2  = openpyxl.Workbook()
                    ws2  = wb2.active
                    ws2.title = "物流單"
                    ws2.append(["收件人", recip_v])
                    ws2.append(["電話",   ph_v])
                    ws2.append(["地址",   addr_v2])
                    ws2.append([])
                    ws2.append(["貨品名稱", "數量"])
                    for it in selected_items:
                        ws2.append([it["name"], it["qty"]])
                    ws2.append([])
                    ws2.append(["合計件數", sum(it["qty"] for it in selected_items)])
                    wb2.save(buf2)
                    buf2.seek(0)
                    fname2 = f"物流單_{recip_v}_{_val(sel_row[waybill_col])}.xlsx"
                    st.download_button(
                        "📥 下載物流單",
                        data=buf2, file_name=fname2,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="logistics_dl")

    # ── Active tracking links ─────────────────────────────────────────────────
    active = [
        (str(row[name_col]), str(row[waybill_col]))
        for _, row in df.iterrows()
        if row[waybill_col]
        and str(row[waybill_col]) not in ("None", "")
        and str(row[status_col]) not in ("None", "")
        and "已簽收" not in str(row[status_col])
        and "已取消" not in str(row[status_col])
    ]

    if active:
        st.divider()
        st.markdown("### 🔗 在途快件 — 快速查詢")
        cols = st.columns(min(len(active), 4))
        for i, (name, wb) in enumerate(active[:12]):
            with cols[i % 4]:
                status_now = str(df.loc[df[waybill_col].astype(str) == wb, status_col].values[0] if len(df.loc[df[waybill_col].astype(str) == wb]) else "")
                badge_color = "#2980b9"
                for kw, (c, _) in _STATUS_COLOR.items():
                    if kw in status_now:
                        badge_color = c; break
                st.markdown(
                    f'<a href="{SF_PUBLIC_TRACK.format(wb)}" target="_blank" style="'
                    f'display:block;background:{badge_color};color:#fff;'
                    f'padding:10px 14px;border-radius:8px;text-decoration:none;'
                    f'margin:4px 0;font-size:13px;line-height:1.5;">'
                    f'<b>{name}</b><br>'
                    f'<span style="font-family:monospace;font-size:11px;opacity:.9">{wb}</span><br>'
                    f'<span style="font-size:11px;opacity:.85">{status_now}</span></a>',
                    unsafe_allow_html=True)

    # ── 小票 PDF 預覽 ─────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 🧾 小票預覽")

    pdf_col  = df.columns[COL_PDF_PATH - 1]
    has_pdf  = df[pdf_col].apply(lambda v: bool(_val(v)) and os.path.exists(_val(v)))
    pdf_rows = df[has_pdf]

    if pdf_rows.empty:
        st.info("暫無小票檔案（完成寄件後自動出現）")
    else:
        options = [
            f"{_val(r[date_col])}  {_val(r[name_col])}  {_val(r[waybill_col])}"
            for _, r in pdf_rows.iterrows()
        ]
        selected = st.selectbox("選擇訂單", options[::-1])   # 最新在最頂
        if selected:
            sel_row  = pdf_rows.iloc[len(options) - 1 - options[::-1].index(selected)]
            pdf_path = _val(sel_row[pdf_col])
            if os.path.exists(pdf_path):
                with open(pdf_path, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                st.markdown(
                    f'<iframe src="data:application/pdf;base64,{b64}" '
                    f'width="100%" height="700" type="application/pdf">'
                    f'</iframe>',
                    unsafe_allow_html=True,
                )
            else:
                st.warning(f"找不到檔案：{pdf_path}")

    # ── 稅金輸入 ──────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### 💰 稅金輸入")
    if _IS_CLOUD:
        st.caption("直接喺稅金欄輸入金額，修改後按「💾 儲存稅金」，自動同步到 GitHub → 本地 Excel")
    else:
        st.caption("直接喺稅金欄輸入金額，修改後按「💾 儲存稅金」")

    tax_data = []
    for _, r in df.iterrows():
        wb_v     = _val(r[waybill_col])
        date_v   = _val(r[date_col])
        name_v   = _val(r[name_col])
        status_v = _val(r[status_col])
        tax_v    = r[tax_col] if tax_col and tax_col in df.columns else None
        try:
            tax_num = float(tax_v) if tax_v not in (None, "", "nan", "None") else None
        except (ValueError, TypeError):
            tax_num = None
        tax_data.append({
            "日期":      date_v,
            "客人":      name_v,
            "運單號":    wb_v,
            "狀態":      status_v,
            "稅金(HKD)": tax_num,
        })

    tax_df = pd.DataFrame(tax_data)
    edited = st.data_editor(
        tax_df,
        column_config={
            "日期":      st.column_config.TextColumn("日期",   disabled=True, width="small"),
            "客人":      st.column_config.TextColumn("客人",   disabled=True, width="small"),
            "運單號":    st.column_config.TextColumn("運單號", disabled=True, width="medium"),
            "狀態":      st.column_config.TextColumn("狀態",   disabled=True, width="small"),
            "稅金(HKD)": st.column_config.NumberColumn(
                "稅金 (HKD)", min_value=0, step=0.1, format="%.1f", width="small",
            ),
        },
        hide_index=True,
        use_container_width=True,
        key="tax_editor",
    )
    if st.button("💾 儲存稅金", type="primary", key="save_tax_btn"):
        saved = _save_tax_values(edited)
        if saved:
            st.success(f"✅ 已儲存 {saved} 筆稅金記錄")
            st.cache_data.clear()
            st.rerun()
        else:
            st.info("沒有可儲存的記錄")

    # ── Auto-refresh ──────────────────────────────────────────────────────────
    st.divider()
    with st.expander("⚙️ 自動更新設定"):
        if st.checkbox("每 60 秒自動重新載入頁面資料（不查詢順豐，只重讀 Excel）"):
            time.sleep(60)
            st.cache_data.clear()
            st.rerun()


if __name__ == "__main__":
    main()
