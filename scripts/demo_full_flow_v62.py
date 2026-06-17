# -*- coding: utf-8 -*-
"""
demo_full_flow_v62.py  ─  完整自動落單流程 V62
=================================================
改用 camp.sf-express.com 批量上傳 Excel 落順丰單，
取代舊版 hk.sf-express.com 表單。

流程：
  ① 彈出輸入視窗，貼入 WhatsApp 訂單
  ② POS 逐張落單，下載小票+清關 PDF
  ③ 生成批量 Excel，上傳到 camp 系統，提交取得運單號
  ④ 掃描打印頁面下載運單 PDF
  ⑤ 整理存檔，更新 tracking.xlsx，git push

執行：python -X utf8 "scripts/demo_full_flow_v62.py"
"""

import os, sys, re, json, time, base64, shutil, subprocess
from datetime import date
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

sys.stdout.reconfigure(encoding="utf-8")

# ─── 路徑 & 常數 ──────────────────────────────────────────────────────────────
CHROME_PROFILE  = r"C:\ChromeAutomation"
ORDERS_DIR      = r"C:\Users\user\Desktop\順丰E順递\data\orders"
PRODUCTS_JSON   = r"C:\Users\user\Desktop\順丰E順递\data\products.json"
EXCEL_PATH      = r"C:\Users\user\Desktop\順丰E順递\data\tracking.xlsx"
BATCH_TEMPLATE  = r"C:\Users\user\Desktop\順丰E順递\批量UPLOAD\寄快遞批量下單模板.xlsx"
TMP_EXCEL       = r"C:\Users\user\Desktop\順丰E順递\批量UPLOAD\_tmp_v62_batch.xlsx"
BATCH_HISTORY   = r"C:\Users\user\Desktop\順丰E順递\批量UPLOAD\歷史記錄"

POS_URL  = "https://online-store-99126206.web.app/"
POS_PASS = "0000"
VIP_PASS = "941196"
CAMP_BATCH_URL  = "https://camp.sf-express.com/web/portal/jkd-hongkong/batchorder"
CAMP_PRINT_URL  = "https://camp.sf-express.com/ScanPrint"

MONTHLY_ACCOUNT = "8526937071"
SENDER = dict(
    name="潘正儀", area=852, mobile="66832382",
    city="香港/Hong Kong", district="黃大仙區",
    region="新蒲崗", address="大有街33號佳力工業大廈603室"
)

today = date.today().strftime("%Y%m%d")
_NUM_ROWS = 15

# ─── 工具函數 ──────────────────────────────────────────────────────────────────

def _load_products():
    try:
        with open(PRODUCTS_JSON, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _parse_order(raw: str) -> dict:
    products = _load_products()
    raw = raw.replace("\n", " ").replace("\r", " ")
    raw = re.sub(r'(?<=[一\-x×\s])([lL])(?=[件個盒包條罐支瓶箱套份粒])', '1', raw)
    raw = re.sub(r'(\d{5,10}[^\d]{0,5})([lL])([件個盒包條罐支瓶箱套份粒])', r'\g<1>1\3', raw)

    items_pos, items_sf = [], []
    ITEM_RE = re.compile(
        r"(\d{5,10})[^\dlL]*?([lL\d]{1,4})\s*[件個盒包條罐支瓶箱套份粒]"
    )
    last_end = 0
    for m in ITEM_RE.finditer(raw):
        sku = m.group(1)
        qty = int(m.group(2).lower().replace('l', '1'))
        last_end = m.end()
        prod = products.get(sku, {})
        items_pos.append({"sku": sku, "qty": qty})
        items_sf.append({
            "sku": sku, "name": prod.get("name", sku),
            "unit_price": float(prod.get("vip_price", 0)), "qty": qty,
        })

    if not items_pos:
        raise ValueError("找唔到貨品 — 請用：型號 數量件/個/盒")

    # 名字解析
    nm = re.search(r"寄\s*([^\s\d，,、]{1,10})\s+", raw)
    if nm:
        name = nm.group(1).strip("，,、 \t")
    else:
        remainder = raw[last_end:].strip().lstrip("，,、—- \t")
        for pat in [
            r"([^\s\d，,、]{1,10})\s+",
            r"([^\d\s，,、]{1,10})(?=1[3-9]\d{9}|[2-9]\d{7})",
        ]:
            nm2 = re.match(pat, remainder)
            if nm2:
                name = nm2.group(1).strip("，,、 \t")
                break
        else:
            nm4 = re.search(r"([^\d\s，,、—\-]{2,10})\s*(?=1[3-9]\d{9}|[2-9]\d{7})", raw)
            if not nm4:
                raise ValueError("找唔到收件人名")
            name = nm4.group(1).strip("，,、 \t")

    m = re.search(r"1[3-9]\d{9}|[2-9]\d{7}", raw)
    if not m:
        raise ValueError("找唔到電話號碼")
    phone = m.group(0)

    m = re.search(r"(?:1[3-9]\d{9}|[2-9]\d{7})\s*(.+)", raw, re.DOTALL)
    if not m:
        raise ValueError("找唔到地址")
    address = re.split(r"[。\n]", m.group(1).strip())[0].strip()

    return {"name": name, "phone": phone, "address": address,
            "items_pos": items_pos, "items_sf": items_sf}


_CITY_ABBREV = {
    "深圳":"深圳市","广州":"广州市","东莞":"东莞市","佛山":"佛山市",
    "珠海":"珠海市","惠州":"惠州市","中山":"中山市","江门":"江门市",
    "北京":"北京市","上海":"上海市","天津":"天津市","重庆":"重庆市",
    "成都":"成都市","武汉":"武汉市","杭州":"杭州市","南京":"南京市",
    "西安":"西安市","苏州":"苏州市","宁波":"宁波市","厦门":"厦门市",
    "福州":"福州市","长沙":"长沙市","郑州":"郑州市","沈阳":"沈阳市",
    "大连":"大连市","济南":"济南市","青岛":"青岛市","合肥":"合肥市",
}

_CITY_PROVINCE = {
    "深圳市":"广东省","广州市":"广东省","东莞市":"广东省","佛山市":"广东省",
    "珠海市":"广东省","惠州市":"广东省","中山市":"广东省","江门市":"广东省",
    "汕头市":"广东省","湛江市":"广东省","肇庆市":"广东省","清远市":"广东省",
    "潮州市":"广东省","梅州市":"广东省","茂名市":"广东省","阳江市":"广东省",
    "河源市":"广东省","云浮市":"广东省","揭州市":"广东省","韶关市":"广东省",
    "北京市":"北京市","上海市":"上海市","天津市":"天津市","重庆市":"重庆市",
    "成都市":"四川省","武汉市":"湖北省","杭州市":"浙江省","南京市":"江苏省",
    "西安市":"陕西省","苏州市":"江苏省","宁波市":"浙江省","厦门市":"福建省",
    "福州市":"福建省","长沙市":"湖南省","郑州市":"河南省","沈阳市":"辽宁省",
    "大连市":"辽宁省","哈尔滨市":"黑龙江省","长春市":"吉林省",
    "济南市":"山东省","青岛市":"山东省","合肥市":"安徽省","石家庄市":"河北省",
    "太原市":"山西省","昆明市":"云南省","贵阳市":"贵州省","南宁市":"广西壮族自治区",
    "海口市":"海南省","兰州市":"甘肃省","无锡市":"江苏省","温州市":"浙江省",
    "南通市":"江苏省","烟台市":"山东省","潍坊市":"山东省","泉州市":"福建省",
}

def _parse_cn_address(addr: str):
    """解析大陸地址 → (province, city, district, detail)"""
    province = re.match(r'([一-鿿]{2,8}(?:省|自治[區区]|自治州|特別行政[區区]|特别行政[區区]))', addr)
    rest = addr[len(province.group(0)):] if province else addr
    province = province.group(0) if province else ''

    city = re.match(r'([一-鿿]{2,8}市)', rest)
    rest = rest[len(city.group(0)):] if city else rest
    city = city.group(0) if city else ''

    district = re.match(r'([一-鿿]{2,8}(?:區|区|縣|县|旗|鎮|镇|鄉|乡|街道))', rest)
    rest = rest[len(district.group(0)):] if district else rest
    district = district.group(0) if district else ''

    # 如果 district 以城市簡稱開頭（如「深圳龙岗区」而非「深圳市龙岗区」），拆開
    if not city and district:
        for abbrev, full_city in _CITY_ABBREV.items():
            if district.startswith(abbrev) and len(district) > len(abbrev):
                city = full_city
                district = district[len(abbrev):]
                break

    # 如果沒有省份但有城市，自動補省份
    if not province and city:
        province = _CITY_PROVINCE.get(city, '')

    return province, city, district, rest.strip()


def _is_hk_phone(phone: str) -> bool:
    return bool(re.match(r'^[2-9]\d{7}$', phone))


# ─── tkinter 輸入視窗 ──────────────────────────────────────────────────────────

def show_order_input() -> list:
    results = [None]
    root = tk.Tk()
    root.title("V62 順丰寄件 — 輸入客人訂單")
    root.resizable(True, True)
    root.geometry("900x900")

    input_lf = tk.LabelFrame(root, text="WhatsApp 訂單 (每行一個客人)", padx=6, pady=6)
    input_lf.pack(fill="x", padx=12, pady=(12, 4))

    row_vars, row_status = [], []
    for i in range(_NUM_ROWS):
        sv  = tk.StringVar()
        rst = tk.StringVar(value="")
        row_vars.append(sv)
        row_status.append(rst)
        frm = tk.Frame(input_lf)
        frm.pack(fill="x", pady=1)
        tk.Label(frm, text=f"{i+1:2d}.", width=3, anchor="e", font=("Courier", 10)).pack(side="left")
        tk.Entry(frm, textvariable=sv, font=("", 10), relief="solid", bd=1).pack(side="left", fill="x", expand=True, padx=4)
        tk.Label(frm, textvariable=rst, width=5, anchor="w", font=("", 9)).pack(side="left")

    result_lf = tk.LabelFrame(root, text="已解析訂單", padx=6, pady=4)
    result_lf.pack(fill="both", expand=True, padx=12, pady=4)
    cols = ("#", "收件人", "電話", "貨品", "狀態")
    tree = ttk.Treeview(result_lf, columns=cols, show="headings", height=8)
    for c, w in zip(cols, [30, 90, 130, 400, 90]):
        tree.heading(c, text=c); tree.column(c, width=w)
    tree.pack(fill="both", expand=True)
    tree.tag_configure("ok", foreground="#27ae60")
    tree.tag_configure("error", foreground="#e74c3c")

    parsed_orders = []
    info_var = tk.StringVar(value="（貼入訂單後按「解析」）")

    def on_parse():
        parsed_orders.clear()
        for item in tree.get_children():
            tree.delete(item)
        ok = 0
        for i, sv in enumerate(row_vars):
            raw = sv.get().strip()
            if not raw:
                row_status[i].set(""); continue
            try:
                o = _parse_order(raw)
                items_str = ", ".join(f"{it['name']}×{it['qty']}" for it in o["items_sf"])
                parsed_orders.append(o)
                tree.insert("", "end", values=(i+1, o["name"], o["phone"], items_str, "待寄"), tags=("ok",))
                row_status[i].set("✅"); ok += 1
            except ValueError as e:
                row_status[i].set("❌")
                tree.insert("", "end", values=(i+1, "—", "—", str(e)[:60], "解析失敗"), tags=("error",))
        run_btn.config(state="normal" if ok else "disabled")
        info_var.set(f"✅ {ok} 個訂單，按「開始」" if ok else "❌ 沒有成功解析的訂單")

    def on_run():
        if parsed_orders:
            results[0] = list(parsed_orders)
            root.destroy()

    tk.Label(root, textvariable=info_var, anchor="w", fg="#555", font=("", 9)).pack(fill="x", padx=14)
    btn_row = tk.Frame(root)
    btn_row.pack(pady=(4, 14))
    tk.Button(btn_row, text="🔍 解析", command=on_parse, bg="#2980b9", fg="white", padx=12, pady=6, font=("", 11)).pack(side="left", padx=6)
    run_btn = tk.Button(btn_row, text="🚀 開始自動化", command=on_run, bg="#27ae60", fg="white", padx=12, pady=6, font=("", 11, "bold"), state="disabled")
    run_btn.pack(side="left", padx=6)
    tk.Button(btn_row, text="✖ 取消", command=lambda: sys.exit(0), bg="#e74c3c", fg="white", padx=12, pady=6, font=("", 11)).pack(side="left", padx=6)
    root.protocol("WM_DELETE_WINDOW", lambda: sys.exit(0))
    root.mainloop()
    return results[0]


# ─── 生成批量 Excel ────────────────────────────────────────────────────────────

def generate_batch_excel(orders_with_ids: list, out_path: str):
    """
    orders_with_ids: [(order_dict, pos_order_no), ...]
    生成可上傳到 camp.sf-express.com 的批量 Excel。
    """
    shutil.copy2(BATCH_TEMPLATE, out_path)
    wb = load_workbook(out_path)
    ws = wb['運單訊息內容 Order Content']

    row = 3
    for order, pos_no in orders_with_ids:
        addr = order["address"]
        phone = order["phone"]
        area_code = 852 if _is_hk_phone(phone) else 86
        province, city, district, detail = _parse_cn_address(addr)

        total_w = round(0.1 * sum(it["qty"] for it in order["items_sf"]), 2)
        order_id = pos_no  # 用POS訂單號作唯一ID

        for i, item in enumerate(order["items_sf"]):
            ws.cell(row, 1, order_id)
            if i == 0:
                # 寄件人
                ws.cell(row, 2,  SENDER["name"])
                ws.cell(row, 3,  SENDER["area"])
                ws.cell(row, 4,  SENDER["mobile"])
                ws.cell(row, 7,  SENDER["city"])
                ws.cell(row, 8,  SENDER["district"])
                ws.cell(row, 9,  SENDER["region"])
                ws.cell(row, 10, SENDER["address"])
                # 收件人
                ws.cell(row, 11, order["name"])
                ws.cell(row, 12, area_code)
                ws.cell(row, 13, phone)
                if province: ws.cell(row, 17, province)
                if city:     ws.cell(row, 18, city)
                if district: ws.cell(row, 19, district)
                ws.cell(row, 20, detail or addr)
                # 包裹
                ws.cell(row, 30, total_w)
                ws.cell(row, 31, 1)
                ws.cell(row, 33, 'E順遞/EC-Ship')
                ws.cell(row, 34, '寄付月結/Pay by Sender (Credit Account)')
                ws.cell(row, 35, '寄付月結/Pay by Sender (Credit Account)')
                ws.cell(row, 36, MONTHLY_ACCOUNT)
            # 物品
            ws.cell(row, 23, item["name"])
            ws.cell(row, 24, 0.1)
            ws.cell(row, 25, item["qty"])
            ws.cell(row, 26, '件/piece')
            ws.cell(row, 27, item["unit_price"])
            ws.cell(row, 28, '港元/HKD')
            ws.cell(row, 29, '中國台灣/Taiwan China')
            row += 1

    wb.save(out_path)
    print(f"  已生成批量 Excel：{out_path}（{row-3} 行）")


# ─── 上傳批量訂單 → 取得運單號 ────────────────────────────────────────────────

def upload_batch_and_get_waybills(page, excel_path: str, orders_with_ids: list) -> dict:
    """
    上傳 Excel，勾選同意，提交，返回 {pos_order_no: waybill_no} dict。
    """
    print("\n▶ 上傳批量訂單...")
    page.goto(CAMP_BATCH_URL, wait_until="networkidle")
    time.sleep(3)
    _dismiss(page)

    # 上傳 Excel
    page.locator("input[type=file]").set_input_files(excel_path)
    time.sleep(6)
    _dismiss(page)

    # 等所有行校驗完成
    print("  等待校驗...")
    for _ in range(30):
        fail_count_el = page.locator("text=導入失敗").first
        if fail_count_el.is_visible(timeout=1000):
            break
        time.sleep(1)
    time.sleep(2)

    # 確認校驗結果
    body = page.inner_text("body")
    fail_match  = re.search(r"導入失敗\s+(\d+)", body)
    ok_match    = re.search(r"校驗成功\s+(\d+)", body)
    fail_n = int(fail_match.group(1)) if fail_match else 0
    ok_n   = int(ok_match.group(1))   if ok_match   else 0
    print(f"  校驗結果：成功 {ok_n} 行，失敗 {fail_n} 行")
    if fail_n > 0:
        page.screenshot(path=os.path.join(ORDERS_DIR, f"_batch_validate_fail_{today}.png"))
        raise RuntimeError(f"Excel 校驗失敗 {fail_n} 行，請查看截圖 _batch_validate_fail_{today}.png")

    # 勾選「我同意」
    page.locator("label", has_text="我同意").click()
    time.sleep(1)

    # 提交訂單，攔截 API 回應
    print("  提交訂單...")
    with page.expect_response(
        lambda r: "createBatchOrder" in r.url and r.status == 200,
        timeout=30000
    ) as resp_info:
        page.locator("button:has-text('提交訂單')").click()

    data = resp_info.value.json()
    time.sleep(3)
    page.screenshot(path=os.path.join(ORDERS_DIR, f"_batch_result_{today}.png"))

    # 原始回應存檔供 debug
    dump_path = os.path.join(ORDERS_DIR, f"_batch_api_{today}.json")
    with open(dump_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  API 回應已存：{dump_path}")

    # 解析運單號：strategy 1 — 從 successRecordList 逐項提取
    waybill_map = {}  # pos_order_no → sf_waybill_no
    order_ids = [pos_no for _, pos_no in orders_with_ids]
    success_list = data.get("successRecordList", [])
    for idx, rec in enumerate(success_list):
        cust_no = rec.get("expressNo", "")
        sf_no = ""

        # 先從 printJson（嵌套 JSON 字串）提取
        print_json_str = rec.get("printJson", "")
        if print_json_str:
            try:
                pj = json.loads(print_json_str)
                sf_no = pj.get("masterWaybillNo", "")
                if not sf_no:
                    sub = pj.get("subWaybillNoList", [])
                    if sub:
                        sf_no = sub[0].get("waybillNo", "")
            except Exception:
                pass

        # 後備：掃描所有欄位值找 SF 格式
        if not sf_no:
            for k, v in rec.items():
                vs = str(v) if v else ""
                if re.match(r'SF\d{10,}', vs):
                    sf_no = vs
                    break

        if not cust_no and idx < len(order_ids):
            cust_no = order_ids[idx]
        if sf_no and cust_no:
            waybill_map[cust_no] = sf_no
        print(f"    [{idx}] {cust_no} → {sf_no}")

    # strategy 2 — 從頁面文字提取 SF 號
    if not waybill_map:
        body = page.inner_text("body")
        sf_nos = list(dict.fromkeys(re.findall(r'SF\d{13,}', body)))
        for i, sf in enumerate(sf_nos):
            if i < len(order_ids):
                waybill_map[order_ids[i]] = sf

    print(f"  運單號：{waybill_map}")
    return waybill_map, data


# ─── 下載運單 PDF ──────────────────────────────────────────────────────────────

def download_waybill_pdf(page, waybill: str, dest_path: str) -> bool:
    """用掃描打印頁面下載運單 PDF。"""
    page.goto(CAMP_PRINT_URL, wait_until="domcontentloaded")
    time.sleep(4)
    _dismiss(page)

    # 選「下載到本地」
    page.locator(".el-select").first.click()
    time.sleep(1)
    page.locator("li.el-select-dropdown__item", has_text="下載到本地").click()
    time.sleep(1)
    _dismiss(page)

    scan_input = page.locator("input[placeholder='此處為掃描結果']")
    scan_input.click()
    scan_input.fill(waybill)
    time.sleep(0.5)

    with page.expect_response(
        lambda r: "eos-scp-core" in r.url and ".pdf" in r.url,
        timeout=25000
    ) as pdf_resp_info:
        page.locator("button", has_text="打印").click()

    pdf_bytes = pdf_resp_info.value.body()
    if len(pdf_bytes) > 1000:
        with open(dest_path, "wb") as f:
            f.write(pdf_bytes)
        return True
    return False


# ─── 通用 dismiss 彈窗 ─────────────────────────────────────────────────────────

def _dismiss(page, max_tries=5):
    for _ in range(max_tries):
        try:
            if page.locator(".el-message-box__wrapper").is_visible(timeout=1500):
                btn = page.locator(".el-message-box__btns button.el-button--primary")
                if btn.is_visible(timeout=1000):
                    btn.click(); time.sleep(0.8); continue
        except Exception:
            pass
        break


# ─── POS：智能填寫函數 ─────────────────────────────────────────────────────────

def shot(page, name):
    pass  # 可選截圖，暫時跳過


# ─── Excel 追蹤表 ──────────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    "日期", "客人名", "POS訂單號", "順丰運單號",
    "收件人", "收件電話", "收件地址",
    "貨品摘要", "件數", "VIP總額(HKD)",
    "付款方式", "運費(HKD)", "最新狀態",
    "狀態更新時間", "異常標記", "小票檔案路徑", "備註", "稅金(HKD)",
]
EXCEL_SHEET = "追蹤表"
_HDR_COL = {h: i+1 for i, h in enumerate(EXCEL_HEADERS)}


def _ensure_excel():
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook(); ws = wb.active; ws.title = EXCEL_SHEET
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, hdr in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.fill = hdr_fill; c.font = hdr_font
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)


def append_order_to_excel(order, waybill, pdf_path, pos_order_no):
    _ensure_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[EXCEL_SHEET]
    items_str = " / ".join(f"{it['name']}×{it['qty']}" for it in order.get("items_sf", []))
    total_amt = sum(it.get("unit_price", 0) * it.get("qty", 1) for it in order.get("items_sf", []))
    total_qty = sum(it.get("qty", 1) for it in order.get("items_sf", []))
    ws.insert_rows(2)
    data = {
        "日期": today, "客人名": order.get("name", ""),
        "POS訂單號": pos_order_no, "順丰運單號": waybill,
        "收件人": order.get("name", ""), "收件電話": order.get("phone", ""),
        "收件地址": order.get("address", ""), "貨品摘要": items_str,
        "件數": total_qty, "VIP總額(HKD)": round(total_amt, 1),
        "付款方式": "月結", "最新狀態": "待更新", "小票檔案路徑": pdf_path,
    }
    for hdr, col in _HDR_COL.items():
        ws.cell(row=2, column=col, value=data.get(hdr, ""))
    wb.save(EXCEL_PATH)
    print(f"  📊 Excel 已記錄: {order.get('name')}  {waybill}")


# ─── Git Push ─────────────────────────────────────────────────────────────────

def git_push_all(orders_done: list):
    base = r"C:\Users\user\Desktop\順丰E順递"
    names = ", ".join(o["name"] for o in orders_done)
    msg   = f"V62 {today}: {names}"
    try:
        subprocess.run(["git", "-C", base, "add", "-A"], check=True, capture_output=True)
        subprocess.run(["git", "-C", base, "commit", "-m", msg], check=True, capture_output=True)
        subprocess.run(["git", "-C", base, "push"], check=True, capture_output=True)
        print(f"  ☁️  已 push: {msg}")
    except subprocess.CalledProcessError as e:
        print(f"  ⚠️  git push 失敗: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════════════════

orders = show_order_input()
if not orders:
    sys.exit(0)

print(f"\n{'='*60}")
print(f"  V62 自動化：共 {len(orders)} 個訂單")
print(f"{'='*60}")

# 檢查 Chrome 是否在跑（不強制 kill，只提示）
chrome_running = bool(
    subprocess.run(["tasklist", "/fi", "imagename eq chrome.exe"],
                   capture_output=True).stdout.count(b"chrome.exe"))
if chrome_running:
    answer = messagebox.askyesno(
        "請關閉 Chrome",
        "偵測到 Chrome 正在執行。\n\n"
        "腳本需要使用 ChromeAutomation Profile。\n"
        "請先關閉所有 Chrome 視窗，然後按「是」繼續。\n\n"
        "（按「否」取消）"
    )
    if not answer:
        sys.exit(0)

# 完成的訂單記錄
completed = []  # [(order, pos_no, waybill, combined_pdf_path)]

with sync_playwright() as pw:
    ctx = pw.chromium.launch_persistent_context(
        CHROME_PROFILE, channel="chrome", headless=False,
        args=["--disable-blink-features=AutomationControlled", "--disable-infobars",
              "--disable-session-crashed-bubble"],
        slow_mo=150, viewport={"width": 1280, "height": 900},
    )

    # ══════════════════════════════════════════════════════════════════════════
    # 階段一：POS 逐張落單
    # ══════════════════════════════════════════════════════════════════════════
    print("\n" + "="*60)
    print("  階段一：POS 落單")
    print("="*60)

    for idx, order in enumerate(orders):
        print(f"\n▶ 訂單 {idx+1}/{len(orders)}: {order['name']}")

        try:
            # 開 POS 頁面
            pos_page = ctx.new_page()
            pos_page.goto(POS_URL, wait_until="domcontentloaded", timeout=20000)
            pos_page.evaluate("""async () => {
                if (navigator.serviceWorker) {
                    const rs = await navigator.serviceWorker.getRegistrations();
                    for (const r of rs) await r.unregister();
                }
                if (window.caches) {
                    const ks = await caches.keys();
                    await Promise.all(ks.map(k => caches.delete(k)));
                }
            }""")
            pos_page.reload(wait_until="domcontentloaded", timeout=20000)
            time.sleep(3)

            # 登入後台
            pos_page.locator("button:has-text('后台管理')").first.click(); time.sleep(0.8)
            pos_page.locator("input[type='password']").first.fill(POS_PASS)
            pos_page.keyboard.press("Enter"); time.sleep(1.5)

            # VIP 價
            pos_page.locator("button:has-text('VIP價')").first.click(); time.sleep(0.8)
            pos_page.locator("input[type='password']").first.fill(VIP_PASS)
            pos_page.keyboard.press("Enter"); time.sleep(1.5)

            # 加貨品
            first_sku = order["items_pos"][0]["sku"]
            try:
                pos_page.wait_for_function(
                    f"() => {{ for(const b of document.querySelectorAll('button')) {{ if(b.textContent.includes('{first_sku}')) return true; }} return false; }}",
                    timeout=20000)
            except Exception:
                pass

            for item in order["items_pos"]:
                try:
                    pos_page.keyboard.press("Escape")
                    btn = pos_page.locator(f"button:has-text('{item['sku']}')").first
                    btn.wait_for(state="visible", timeout=15000)
                    btn.scroll_into_view_if_needed(timeout=5000)
                    time.sleep(0.3)
                    for _ in range(item["qty"]):
                        btn.click(); time.sleep(0.25)
                except Exception as e:
                    print(f"  [跳過] SKU {item['sku']}: {e}")

            # 結帳
            pos_page.locator("button:has-text('結帳')").first.click()
            _confirm_btn = pos_page.locator("button:has-text('確認，出小票')").first
            _confirm_btn.wait_for(state="visible", timeout=15000)
            time.sleep(2)
            pos_page.keyboard.press("Escape"); time.sleep(0.3)
            pos_page.evaluate("""() => {
                document.querySelectorAll('*').forEach(el => {
                    if (el.scrollHeight > el.clientHeight + 5 &&
                        getComputedStyle(el).overflowY !== 'visible' &&
                        el.clientHeight > 50 && el.clientHeight < 500) {
                        el.scrollTop = el.scrollHeight;
                    }
                });
            }"""); time.sleep(0.5)
            pos_page.bring_to_front(); time.sleep(0.3)

            for mname, mfn in [
                ("force",    lambda: _confirm_btn.click(force=True)),
                ("dispatch", lambda: _confirm_btn.dispatch_event('click')),
                ("tap",      lambda: _confirm_btn.tap(force=True)),
            ]:
                try: mfn()
                except: pass
                time.sleep(1.5)
                if not pos_page.locator("button:has-text('確認，出小票')").is_visible(timeout=500):
                    print(f"  ✅ 確認按鈕成功（{mname}）"); break
            time.sleep(2)

            # 取 POS 訂單號
            body_text = pos_page.inner_text("body")
            m = re.search(r"ORD-\d+", body_text)
            pos_order_no = m.group(0) if m else f"ORD-{today}{idx}"
            print(f"  ✅ POS 訂單：{pos_order_no}")

            # 下載合一 PDF
            order_folder = f"{order['name']}_{today}_{pos_order_no}"
            save_dir = os.path.join(ORDERS_DIR, order_folder)
            os.makedirs(save_dir, exist_ok=True)
            file_base = f"{order['name']}_{today}_{pos_order_no}"
            combined_path = os.path.join(save_dir, f"{file_base}_明細+清關.pdf")

            print("  下載合一 PDF...")
            try:
                pos_page.wait_for_function("""() => {
                    for (const a of document.querySelectorAll('a[download]')) {
                        if ((a.getAttribute('download')||'').includes('明細') &&
                            (a.getAttribute('href')||'').startsWith('blob:')) return true;
                    }
                    return false;
                }""", timeout=120000)
                time.sleep(5)
                with pos_page.expect_download(timeout=30000) as dl_info:
                    pos_page.evaluate("""() => {
                        for (const a of document.querySelectorAll('a[download]')) {
                            if ((a.getAttribute('download') || '').includes('明細')) {
                                a.click(); return;
                            }
                        }
                    }""")
                dl_info.value.save_as(combined_path)
                print(f"  ✅ PDF 已儲存：{combined_path}")

                # 重排頁序（報關→第1，小票→第2）
                try:
                    import fitz as _fitz
                    _doc = _fitz.open(combined_path)
                    if len(_doc) >= 3:
                        _new = _fitz.open()
                        _new.insert_pdf(_doc, from_page=2, to_page=2)
                        _new.insert_pdf(_doc, from_page=0, to_page=0)
                        _doc.close()
                        _new.save(combined_path, garbage=4, deflate=True)
                        _new.close()
                except Exception:
                    pass

            except Exception as e:
                print(f"  ⚠️  PDF 下載失敗：{e}")
                combined_path = ""

            # 關閉 POS 頁面
            try:
                done = pos_page.locator("button:has-text('完成')").first
                if done.is_visible(timeout=2000): done.click(); time.sleep(0.8)
            except Exception:
                pass
            pos_page.close()

            completed.append({
                "order": order, "pos_no": pos_order_no,
                "combined_pdf": combined_path, "save_dir": save_dir,
                "waybill": None, "waybill_pdf": None,
            })

        except Exception as e:
            print(f"  ❌ POS 失敗：{e}")
            pos_page.close()

    if not completed:
        print("所有 POS 訂單失敗，退出")
        ctx.close()
        sys.exit(1)

    # ══════════════════════════════════════════════════════════════════════════
    # 階段二：生成 Excel 批量上傳順丰
    # ══════════════════════════════════════════════════════════════════════════
    print("\n" + "="*60)
    print("  階段二：SF Camp 批量上傳")
    print("="*60)

    orders_with_ids = [(c["order"], c["pos_no"]) for c in completed]
    generate_batch_excel(orders_with_ids, TMP_EXCEL)

    sf_page = ctx.new_page()
    waybill_map, raw_resp = upload_batch_and_get_waybills(sf_page, TMP_EXCEL, orders_with_ids)

    # 如果 API 沒有直接返回 waybill，嘗試從頁面文字提取
    if not waybill_map:
        print(f"  ⚠️  API 未返回運單號，原始回應：{str(raw_resp)[:300]}")
        print("  請查看截圖確認是否成功，手動記錄運單號")

    # 儲存歷史 Excel（含日期+客人名）
    os.makedirs(BATCH_HISTORY, exist_ok=True)
    names_str = "+".join(c["order"]["name"] for c in completed[:4])
    if len(completed) > 4:
        names_str += f"+等{len(completed)}人"
    history_excel = os.path.join(BATCH_HISTORY, f"批量上傳_{today}_{names_str}.xlsx")
    shutil.copy2(TMP_EXCEL, history_excel)
    print(f"  📋 Excel 已存檔：{os.path.basename(history_excel)}")

    # 配對 waybill 到 completed，並重命名資料夾+PDF（加入 SF 號）
    for c in completed:
        c["waybill"] = waybill_map.get(c["pos_no"], "")
        waybill = c["waybill"]
        if not waybill:
            continue
        name = c["order"]["name"]
        old_dir = c["save_dir"]
        new_dir = os.path.join(ORDERS_DIR, f"{name}_{today}_{c['pos_no']}_{waybill}")
        if old_dir != new_dir and os.path.exists(old_dir):
            os.makedirs(new_dir, exist_ok=True)
            for f in os.listdir(old_dir):
                old_f = os.path.join(old_dir, f)
                # 統一命名加入 SF 號
                if "_明細+清關" in f:
                    new_f = os.path.join(new_dir, f"{name}_{today}_{c['pos_no']}_{waybill}_明細+清關.pdf")
                else:
                    new_f = os.path.join(new_dir, f)
                moved = False
                for attempt in range(10):
                    try:
                        shutil.move(old_f, new_f); moved = True; break
                    except PermissionError:
                        time.sleep(1.5)
                if not moved:
                    shutil.copy2(old_f, new_f)
                    for attempt in range(10):
                        try:
                            os.remove(old_f); break
                        except PermissionError:
                            time.sleep(1.5)
            for attempt in range(10):
                try:
                    shutil.rmtree(old_dir); break
                except Exception:
                    time.sleep(1.5)
            c["save_dir"] = new_dir
            c["combined_pdf"] = os.path.join(new_dir, f"{name}_{today}_{c['pos_no']}_{waybill}_明細+清關.pdf")
            print(f"  📁 資料夾更新：{os.path.basename(new_dir)}")

    # ══════════════════════════════════════════════════════════════════════════
    # 階段三：下載運單 PDF
    # ══════════════════════════════════════════════════════════════════════════
    print("\n" + "="*60)
    print("  階段三：下載運單 PDF")
    print("="*60)

    print("  選擇「下載到本地」...")
    sf_page.goto(CAMP_PRINT_URL, wait_until="domcontentloaded")
    time.sleep(4)
    _dismiss(sf_page)
    sf_page.locator(".el-select").first.click(); time.sleep(1)
    sf_page.locator("li.el-select-dropdown__item", has_text="下載到本地").click()
    time.sleep(1)
    _dismiss(sf_page)

    scan_input = sf_page.locator("input[placeholder='此處為掃描結果']")

    for c in completed:
        waybill = c.get("waybill", "")
        if not waybill:
            print(f"  ⚠️  {c['order']['name']} 無運單號，跳過下載")
            continue

        dest_path = os.path.join(c["save_dir"],
                                 f"{c['order']['name']}_{today}_{c['pos_no']}_{waybill}_運單.pdf")
        print(f"  下載 {waybill} ({c['order']['name']})...")

        _dismiss(sf_page)
        scan_input.click()
        scan_input.fill(waybill)
        time.sleep(0.5)

        try:
            with sf_page.expect_response(
                lambda r: "eos-scp-core" in r.url and ".pdf" in r.url,
                timeout=25000
            ) as pdf_resp_info:
                sf_page.locator("button", has_text="打印").click()

            pdf_bytes = pdf_resp_info.value.body()
            if len(pdf_bytes) > 1000:
                with open(dest_path, "wb") as f:
                    f.write(pdf_bytes)
                c["waybill_pdf"] = dest_path
                print(f"  ✅ 已儲存：{dest_path}")
            else:
                print(f"  ⚠️  PDF 太小")
        except Exception as e:
            print(f"  ⚠️  下載失敗：{e}")

        time.sleep(1.5)

    sf_page.close()

    # ══════════════════════════════════════════════════════════════════════════
    # 階段四：報關上傳（hk.sf-express.com）
    # ══════════════════════════════════════════════════════════════════════════
    print("\n" + "="*60)
    print("  階段四：報關上傳")
    print("="*60)

    # 引入 clearance_upload 的函數
    import importlib.util as _ilu
    _cl_path = os.path.join(os.path.dirname(__file__), "clearance_upload.py")
    _cl_spec = _ilu.spec_from_file_location("clearance_upload", _cl_path)
    _cl_mod  = _ilu.module_from_spec(_cl_spec)
    _cl_spec.loader.exec_module(_cl_mod)

    CLEARANCE_URL  = "https://hk.sf-express.com/hk/tc/clearance?type=upload"
    IFRAME_HOST    = "sf-international.com"
    SESSION_FILE   = r"C:\Users\user\Desktop\順丰E順递\data\last_session.json"

    cl_page = ctx.new_page()
    sessions_out = []

    for c in completed:
        if not c.get("waybill"):
            continue
        name        = c["order"]["name"]
        waybill     = c["waybill"]
        combined    = c.get("combined_pdf", "")

        front, back = _cl_mod.find_id_cards(name)
        if not front or not os.path.exists(str(front)):
            print(f"  ⚠️  {name} 找不到身份証正面，跳過報關")
            continue
        if not back or not os.path.exists(str(back)):
            print(f"  ⚠️  {name} 找不到身份証背面，跳過報關")
            continue
        if not combined or not os.path.exists(combined):
            print(f"  ⚠️  {name} 找不到清關PDF，跳過報關")
            continue

        print(f"\n▶ 報關 {name}  {waybill}")
        try:
            frame = _cl_mod._get_iframe(cl_page)
            frame.locator("input.ant-input").first.fill(waybill)
            frame.locator("input.ant-input").first.press("Tab")
            time.sleep(1)
            print(f"  運單號已填入：{waybill}")

            _cl_mod._do_id_tab(cl_page, frame, front, back, print)
            _cl_mod._do_customs_tab(cl_page, frame, combined, combined, print)
            print(f"  ✅ {name} 報關完成")

            sessions_out.append({
                "customer": name, "waybill": waybill,
                "pdf_path": c.get("waybill_pdf", ""),
                "id_uploaded": True, "customs_uploaded": True,
            })
        except Exception as e:
            print(f"  ⚠️  {name} 報關失敗：{e}")
            sessions_out.append({
                "customer": name, "waybill": waybill,
                "pdf_path": c.get("waybill_pdf", ""),
                "id_uploaded": False, "customs_uploaded": False,
            })

    # 寫 last_session.json 供補跑用
    if sessions_out:
        with open(SESSION_FILE, "w", encoding="utf-8") as f:
            json.dump(sessions_out, f, ensure_ascii=False, indent=2)
        print(f"\n  已寫入 last_session.json（{len(sessions_out)} 筆）")

    cl_page.close()
    ctx.close()

# ══════════════════════════════════════════════════════════════════════════════
# 階段五：記錄 + Git Push
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  階段五：記錄 + Git Push")
print("="*60)

for c in completed:
    if c.get("waybill"):
        pdf_rel = c["combined_pdf"].replace(r"C:\Users\user\Desktop\順丰E順递" + "\\", "")
        append_order_to_excel(c["order"], c["waybill"], pdf_rel, c["pos_no"])

orders_done = [c["order"] for c in completed if c.get("waybill")]
if orders_done:
    git_push_all(orders_done)

# 最終報告
print("\n" + "="*60)
print("  完成！")
print("="*60)
for c in completed:
    status = "✅" if c.get("waybill") else "⚠️ 無運單"
    print(f"  {status} {c['order']['name']:12s}  POS:{c['pos_no']}  SF:{c.get('waybill','')}")
    if c.get("combined_pdf"):
        print(f"     小票：{os.path.basename(c['combined_pdf'])}")
    if c.get("waybill_pdf"):
        print(f"     運單：{os.path.basename(c['waybill_pdf'])}")
print("="*60)
